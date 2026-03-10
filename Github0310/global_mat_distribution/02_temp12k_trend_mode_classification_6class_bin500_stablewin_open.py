# -*- coding: utf-8 -*-
"""
Temp12k classification (stable windows / 500-year binning / strict zero rule) — 6-class output
====================================================

Input: Temp12k master table in "TRUE_withCompositePriority" format (multiple sheets, each series stored as an Age+Value two-column block)
Binning: 500-year bins -> take the mean in each bin -> perform linear regression on the bin-mean series within each window to calculate slope
Rule: strict zero threshold
    slope < 0  -> W
    slope >= 0 -> C

Stable windows (all windows are fixed to reduce boundary effects/noise):
    overall : 1–12k  (1000–12000)
    early   : 7–11k  (7000–11000)
    mid     : 4–8k   (4000–8000)
    late    : 1–5k   (1000–5000)
    half1   : 1–6k   (1000–6000)   # extra WCW constraint / half-window split detection
    half2   : 6–12k  (6000–12000)

Classification (mutually exclusive; stop at the first match in order):
1) WCW：
   - overall(1–12k) is W
   - pattern(7–11,4–8,1–5) == "WCW"
   - half1(1–6k) is W and half2(6–12k) is W (both segments have slope < 0 and enough bins)
   Once WCW is assigned, do not enter later filters.

2) MHM：
   - (overall is C but pattern is WCC/WWC) OR (overall is W but pattern is WCC/WWC)
   - Additionally, if the half-window signs split (half1=C and half2=W), assign directly to MHM
     (to avoid cases like "0-6>0, 6-12<0" being scattered into other classes)

3) CCC： pattern == "CCC"
4) WWW： pattern == "WWW"
5) MIX：all remaining valid series
6) INSUF：unable to calculate overall or any of the three pattern slopes (insufficient bins / NaN / no data)

Output:
out_root/
  ├─ WCW/
  ├─ MHM/
  ├─ CCC/
  ├─ WWW/
  ├─ MIX/
  └─ INSUF/
Each class contains one *_CLASS_*.xlsx file (preserving the original sheet structure and writing back by block)
Also generate one SUMMARY.xlsx file (all slopes / window bin counts / class for each series)

----------------------------------------------------
Note:
- For manually added western-longitude sites: Cartopy longitudes are generally expressed in [-180, 180], so negative values are preferred for western longitudes.
"""

from __future__ import annotations

import os
from pathlib import Path
import re
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional

import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font

# ===================== Edit here only =====================
SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_XLSX = str(SCRIPT_DIR / "2026-01-14_1426_Temp12k_Temperature_TRUE_withCompositePriority.xlsx")
OUTPUT_BASE_DIR = str(SCRIPT_DIR)  # Output to the script directory by default
# =====================================================

# -------------------- Binning and windows --------------------
BIN_WIDTH_Y = 500.0

AGE_LO = 0.0
AGE_HI = 12000.0

# Stable windows (unit: year BP)
WIN_OVERALL = (1000.0, 12000.0)  # 1–12k
WIN_EARLY = (7000.0, 11000.0)    # 7–11k
WIN_MID = (4000.0, 8000.0)       # 4–8k
WIN_LATE = (1000.0, 5000.0)      # 1–5k

WIN_HALF1 = (1000.0, 6000.0)     # 1–6k
WIN_HALF2 = (6000.0, 12000.0)    # 6–12k

# Minimum required number of valid bins
MIN_BINS_OVERALL = 6
MIN_BINS_SEG = 3
MIN_BINS_HALF = 6

# -------------------- Output classes --------------------
CLASS_ORDER = ["WCW", "MHM", "CCC", "WWW", "MIX", "INSUF"]

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BOLD = Font(bold=True)

EXCEL_MAX_COLS = 16384
BAD_SHEET_CHARS = [":", "\\", "/", "?", "*", "[", "]"]

LONLAT_RE = re.compile(r"Lon\s*=\s*([-+]?\d*\.?\d+)\s*,\s*Lat\s*=\s*([-+]?\d*\.?\d+)", re.I)


# -------------------- Basic utilities --------------------
def to_long_path(path: str) -> str:
    p = os.path.abspath(path)
    if os.name == "nt":
        if p.startswith("\\\\?\\"):
            return p
        if len(p) > 240:
            return "\\\\?\\" + p
    return p


def safe_short_name(path_or_name: str, maxlen: int = 40) -> str:
    s = os.path.splitext(os.path.basename(path_or_name))[0]
    s = re.sub(r"[^\w\-]+", "_", s).strip("_")
    return (s[:maxlen] or "OUT").strip("_")


def safe_sheet_name(name: str) -> str:
    s = str(name).strip()
    for ch in BAD_SHEET_CHARS:
        s = s.replace(ch, "_")
    return (s or "sheet")[:31]


def is_age_header(v) -> bool:
    return str(v).strip().lower() == "age"


def is_value_header(v) -> bool:
    return str(v).strip().lower() == "value"


def detect_header_row(ws, row_min: int = 3, row_max: int = 10) -> int:
    """
    Some headers may not be on row 4, so search rows [3..10] and use the row with the most "Age/Value" pairs as header_row
    """
    best_row = 4
    best_pairs = -1

    for r in range(row_min, row_max + 1):
        pairs = 0
        for c in range(1, ws.max_column):
            v1 = ws.cell(r, c).value
            v2 = ws.cell(r, c + 1).value
            if is_age_header(v1) and is_value_header(v2):
                pairs += 1
        if pairs > best_pairs:
            best_pairs = pairs
            best_row = r

    return best_row


def iter_blocks(ws) -> Tuple[List[int], int]:
    """
    Return:
      starts: starting column of each series block (the Age column)
      header_row: row containing Age/Value
    """
    header_row = detect_header_row(ws)
    starts: List[int] = []
    for c in range(1, ws.max_column):
        v1 = ws.cell(header_row, c).value
        v2 = ws.cell(header_row, c + 1).value
        if is_age_header(v1) and is_value_header(v2):
            starts.append(c)
    return starts, header_row


def parse_lon_lat(coord_str: str) -> Tuple[str, str]:
    if not coord_str:
        return "", ""
    m = LONLAT_RE.search(str(coord_str))
    if not m:
        return "", ""
    return m.group(1), m.group(2)


def maybe_ka_to_years(ages: np.ndarray) -> np.ndarray:
    """
    If Age max <= 200, treat it as ka BP and convert to years BP
    """
    if ages.size == 0:
        return ages
    mx = np.nanmax(ages)
    if np.isfinite(mx) and mx <= 200.0:
        return ages * 1000.0
    return ages


def extract_series(ws, c: int, header_row: int) -> Tuple[List[float], List[float]]:
    """
    Extract numeric values from a series block:
    - header_row contains Age/Value
    - data start from header_row + 1
    - collect values whenever both Age and Value can be converted to float
    """
    ages: List[float] = []
    vals: List[float] = []

    # Find the last valid row in these two columns to reduce unnecessary looping
    max_row = ws.max_row
    last = header_row
    for r in range(max_row, header_row, -1):
        if ws.cell(r, c).value is not None or ws.cell(r, c + 1).value is not None:
            last = r
            break
    if last <= header_row:
        return ages, vals

    for r in range(header_row + 1, last + 1):
        a = ws.cell(r, c).value
        v = ws.cell(r, c + 1).value
        if a is None or v is None:
            continue
        try:
            af = float(a)
            vf = float(v)
        except Exception:
            continue
        if not (np.isfinite(af) and np.isfinite(vf)):
            continue
        ages.append(af)
        vals.append(vf)

    return ages, vals


# -------------------- 500-year binned mean --------------------
def bin_mean_series(
    ages: List[float],
    vals: List[float],
    lo: float = AGE_LO,
    hi: float = AGE_HI,
    bin_width: float = BIN_WIDTH_Y,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Return (centers, means):
    centers: bin centers
    means: mean of each bin (NaN if no data)
    """
    x = np.asarray(ages, dtype=float)
    y = np.asarray(vals, dtype=float)
    x = maybe_ka_to_years(x)

    m = np.isfinite(x) & np.isfinite(y) & (x >= lo) & (x <= hi)
    x = x[m]
    y = y[m]

    edges = np.arange(lo, hi + bin_width, bin_width)
    centers = edges[:-1] + bin_width / 2.0
    means = np.full_like(centers, np.nan, dtype=float)

    if x.size == 0:
        return centers, means

    for i in range(len(centers)):
        a0 = edges[i]
        a1 = edges[i + 1]
        if i < len(centers) - 1:
            mm = (x >= a0) & (x < a1)
        else:
            mm = (x >= a0) & (x <= a1)
        if np.any(mm):
            means[i] = float(np.mean(y[mm]))

    return centers, means


def slope_from_binned(
    centers: np.ndarray,
    means: np.ndarray,
    lo: float,
    hi: float,
    min_bins: int,
) -> Tuple[float, int]:
    """
    Perform linear regression on bin means within [lo, hi] to compute slope (unit: value/year)
    """
    x = np.asarray(centers, dtype=float)
    y = np.asarray(means, dtype=float)

    m = np.isfinite(x) & np.isfinite(y) & (x >= lo) & (x <= hi)
    xw = x[m]
    yw = y[m]
    n = int(xw.size)
    if n < min_bins:
        return float("nan"), n

    xbar = float(np.mean(xw))
    ybar = float(np.mean(yw))
    denom = float(np.sum((xw - xbar) ** 2))
    if denom == 0.0:
        return float("nan"), n

    numer = float(np.sum((xw - xbar) * (yw - ybar)))
    return numer / denom, n


def wc_letter_strict0(slope: float) -> str:
    if not np.isfinite(slope):
        return ""
    return "W" if slope < 0.0 else "C"


# -------------------- Write output blocks (preserve the original table structure) --------------------
def write_block(
    out_ws,
    start_col: int,
    site: str,
    coord: str,
    unit: str,
    ages: List[float],
    vals: List[float],
):
    c1, c2 = start_col, start_col + 1

    out_ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    out_ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
    out_ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)

    cell = out_ws.cell(1, c1, site)
    cell.font = BOLD
    cell.alignment = CENTER

    cell = out_ws.cell(2, c1, coord)
    cell.alignment = CENTER

    cell = out_ws.cell(3, c1, unit)
    cell.alignment = CENTER

    out_ws.cell(4, c1, "Age").font = BOLD
    out_ws.cell(4, c2, "Value").font = BOLD

    n = min(len(ages), len(vals))
    for i in range(n):
        out_ws.cell(5 + i, c1, float(ages[i]))
        out_ws.cell(5 + i, c2, float(vals[i]))


def ensure_out_wb_like_input(wb_in) -> Tuple[Workbook, Dict[str, Any], Dict[str, int]]:
    out = Workbook()
    out.remove(out.active)

    sheet_map: Dict[str, Any] = {}
    next_col: Dict[str, int] = {}

    for sheet_name in wb_in.sheetnames:
        ws = out.create_sheet(safe_sheet_name(sheet_name))
        sheet_map[sheet_name] = ws
        next_col[sheet_name] = 1

    return out, sheet_map, next_col


# -------------------- Classification logic (mutually exclusive; stop at the first match) --------------------
def decide_class_6way(
    overall_letter: str,
    pattern3: str,
    s_half1: float,
    s_half2: float,
    n_half1: int,
    n_half2: int,
) -> str:
    """
    Classification order: WCW -> MHM -> CCC -> WWW -> MIX
    If calculation requirements are not met (missing overall or pattern), assign directly to INSUF in the outer logic
    """

    # ========== 1) WCW (strictest priority) ==========
    if overall_letter == "W" and pattern3 == "WCW":
        ok_half = (
            np.isfinite(s_half1) and np.isfinite(s_half2)
            and (n_half1 >= MIN_BINS_HALF) and (n_half2 >= MIN_BINS_HALF)
            and (s_half1 < 0.0) and (s_half2 < 0.0)
        )
        if ok_half:
            return "WCW"
        # If pattern=WCW but the half-window condition is not met, do not discard it; continue to later checks (it may go to MHM or MIX)

    # ========== 2) MHM ==========
    # 2A/2B: if pattern is WCC or WWC, assign it regardless of whether overall is W or C
    if pattern3 in {"WCC", "WWC"}:
        return "MHM"

    # 2C: half-window sign split: half1=C and half2=W (to avoid scattering into other classes)
    # Note: half-window slopes may be unavailable; they must be finite and have enough bins
    if (
        np.isfinite(s_half1) and np.isfinite(s_half2)
        and (n_half1 >= MIN_BINS_HALF) and (n_half2 >= MIN_BINS_HALF)
    ):
        half1_letter = wc_letter_strict0(s_half1)  # W/C
        half2_letter = wc_letter_strict0(s_half2)
        if half1_letter == "C" and half2_letter == "W":
            return "MHM"

    # ========== 3) CCC ==========
    if pattern3 == "CCC":
        return "CCC"

    # ========== 4) WWW ==========
    if pattern3 == "WWW":
        return "WWW"

    # ========== 5) Remaining valid series -> MIX ==========
    return "MIX"


# -------------------- SUMMARY output --------------------
def build_summary_workbook(
    out_path: str,
    params: Dict[str, Any],
    rows: List[Dict[str, Any]],
    class_counts: Dict[str, int],
):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "COUNTS"

    ws0["A1"] = "Class"
    ws0["B1"] = "Count"
    ws0["C1"] = "Percent(valid)"
    ws0["A1"].font = BOLD
    ws0["B1"].font = BOLD
    ws0["C1"].font = BOLD

    valid_total = sum(class_counts.get(k, 0) for k in CLASS_ORDER if k != "INSUF")
    r = 2
    for k in CLASS_ORDER:
        ws0.cell(r, 1, k)
        ws0.cell(r, 2, int(class_counts.get(k, 0)))
        if k != "INSUF" and valid_total > 0:
            ws0.cell(r, 3, float(class_counts.get(k, 0)) / float(valid_total))
        else:
            ws0.cell(r, 3, "")
        r += 1

    wsP = wb.create_sheet("PARAMS")
    wsP["A1"] = "Key"
    wsP["B1"] = "Value"
    wsP["A1"].font = BOLD
    wsP["B1"].font = BOLD
    rr = 2
    for k, v in params.items():
        wsP.cell(rr, 1, str(k))
        wsP.cell(rr, 2, str(v))
        rr += 1

    ws = wb.create_sheet("SERIES")
    headers = [
        "final_class",
        "overall_letter(1-12k)",
        "pattern3(7-11|4-8|1-5)",
        "sheet",
        "site",
        "lon",
        "lat",
        "unit",
        "n_bins_overall(1-12k)",
        "slope_1_12k",
        "n_bins_early(7-11k)",
        "slope_7_11k",
        "n_bins_mid(4-8k)",
        "slope_4_8k",
        "n_bins_late(1-5k)",
        "slope_1_5k",
        "n_bins_half1(1-6k)",
        "slope_1_6k",
        "n_bins_half2(6-12k)",
        "slope_6_12k",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h)
        c.font = BOLD

    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            ws.cell(i, j, row.get(h, ""))

    wb.save(to_long_path(out_path))


# -------------------- Main program --------------------
def main():
    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(f"Input file not found: {INPUT_XLSX}")

    wb_in = load_workbook(INPUT_XLSX, data_only=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    short_base = safe_short_name(INPUT_XLSX, maxlen=35)

    root_base = OUTPUT_BASE_DIR if OUTPUT_BASE_DIR else os.getcwd()
    out_root = os.path.join(root_base, f"{short_base}_6class_500bin_stablewin_strict0_{ts}")
    os.makedirs(out_root, exist_ok=True)

    # 6-class output directories + workbooks
    class_dirs: Dict[str, str] = {}
    out_wbs: Dict[str, Workbook] = {}
    out_sheets: Dict[str, Dict[str, Any]] = {}
    out_nextcol: Dict[str, Dict[str, int]] = {}

    for cls in CLASS_ORDER:
        d = os.path.join(out_root, cls)
        os.makedirs(d, exist_ok=True)
        class_dirs[cls] = d

        wb_out, sheet_map, next_col = ensure_out_wb_like_input(wb_in)
        out_wbs[cls] = wb_out
        out_sheets[cls] = sheet_map
        out_nextcol[cls] = next_col

    total_blocks = 0
    class_counts: Dict[str, int] = {k: 0 for k in CLASS_ORDER}
    summary_rows: List[Dict[str, Any]] = []

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        starts, header_row = iter_blocks(ws)

        # Do not raise an error if no blocks are detected in this sheet (it may be an empty sheet)
        for c in starts:
            # Read header information first
            site = ws.cell(1, c).value
            coord = ws.cell(2, c).value
            unit = ws.cell(3, c).value
            site = "" if site is None else str(site)
            coord = "" if coord is None else str(coord)
            unit = "" if unit is None else str(unit)

            # Extract data (even empty series should still enter INSUF)
            ages, vals = extract_series(ws, c, header_row=header_row)

            # Default class: INSUF
            final_cls = "INSUF"

            # Binning and slopes (if there is no data, all values are NaN)
            centers, means = bin_mean_series(ages, vals, lo=AGE_LO, hi=AGE_HI, bin_width=BIN_WIDTH_Y)

            s_overall, n_overall = slope_from_binned(centers, means, WIN_OVERALL[0], WIN_OVERALL[1], MIN_BINS_OVERALL)
            s_e, n_e = slope_from_binned(centers, means, WIN_EARLY[0], WIN_EARLY[1], MIN_BINS_SEG)
            s_m, n_m = slope_from_binned(centers, means, WIN_MID[0], WIN_MID[1], MIN_BINS_SEG)
            s_l, n_l = slope_from_binned(centers, means, WIN_LATE[0], WIN_LATE[1], MIN_BINS_SEG)

            # Half-window slopes (not used for INSUF classification; only used for WCW / MHM split checks)
            s_h1, n_h1 = slope_from_binned(centers, means, WIN_HALF1[0], WIN_HALF1[1], MIN_BINS_HALF)
            s_h2, n_h2 = slope_from_binned(centers, means, WIN_HALF2[0], WIN_HALF2[1], MIN_BINS_HALF)

            overall_letter = ""
            pattern3 = ""

            # ✅ Valid-series criterion: overall and all three pattern slopes must be computable
            if np.isfinite(s_overall) and np.isfinite(s_e) and np.isfinite(s_m) and np.isfinite(s_l):
                overall_letter = wc_letter_strict0(s_overall)
                pattern3 = wc_letter_strict0(s_e) + wc_letter_strict0(s_m) + wc_letter_strict0(s_l)

                final_cls = decide_class_6way(
                    overall_letter=overall_letter,
                    pattern3=pattern3,
                    s_half1=s_h1,
                    s_half2=s_h2,
                    n_half1=n_h1,
                    n_half2=n_h2,
                )
            else:
                final_cls = "INSUF"

            # Write to the corresponding class workbook (no blocks are dropped)
            out_ws = out_sheets[final_cls][sheet_name]
            col = out_nextcol[final_cls][sheet_name]
            if col + 1 > EXCEL_MAX_COLS:
                raise RuntimeError(f"Sheet '{sheet_name}' in class '{final_cls}' exceeds the Excel column limit.")

            write_block(out_ws, col, site, coord, unit, ages, vals)
            out_nextcol[final_cls][sheet_name] += 2

            class_counts[final_cls] += 1
            total_blocks += 1

            lon_s, lat_s = parse_lon_lat(coord)

            summary_rows.append({
                "final_class": final_cls,
                "overall_letter(1-12k)": overall_letter,
                "pattern3(7-11|4-8|1-5)": pattern3,
                "sheet": sheet_name,
                "site": site,
                "lon": lon_s,
                "lat": lat_s,
                "unit": unit,
                "n_bins_overall(1-12k)": int(n_overall),
                "slope_1_12k": float(s_overall) if np.isfinite(s_overall) else "",
                "n_bins_early(7-11k)": int(n_e),
                "slope_7_11k": float(s_e) if np.isfinite(s_e) else "",
                "n_bins_mid(4-8k)": int(n_m),
                "slope_4_8k": float(s_m) if np.isfinite(s_m) else "",
                "n_bins_late(1-5k)": int(n_l),
                "slope_1_5k": float(s_l) if np.isfinite(s_l) else "",
                "n_bins_half1(1-6k)": int(n_h1),
                "slope_1_6k": float(s_h1) if np.isfinite(s_h1) else "",
                "n_bins_half2(6-12k)": int(n_h2),
                "slope_6_12k": float(s_h2) if np.isfinite(s_h2) else "",
            })

    # Save xlsx files for all classes
    for cls in CLASS_ORDER:
        out_name = f"{short_base}_{cls}_{ts}.xlsx"
        out_path = os.path.join(class_dirs[cls], out_name)
        out_wbs[cls].save(to_long_path(out_path))

    # summary
    params = {
        "INPUT_XLSX": INPUT_XLSX,
        "BIN_WIDTH_Y": BIN_WIDTH_Y,
        "AGE_RANGE_USED_FOR_BINNING": f"{AGE_LO}-{AGE_HI}",
        "STRICT_RULE": "W if slope<0 else C (>=0)",
        "WINDOW_OVERALL": str(WIN_OVERALL),
        "WINDOW_EARLY": str(WIN_EARLY),
        "WINDOW_MID": str(WIN_MID),
        "WINDOW_LATE": str(WIN_LATE),
        "WINDOW_HALF1": str(WIN_HALF1),
        "WINDOW_HALF2": str(WIN_HALF2),
        "MIN_BINS_OVERALL": MIN_BINS_OVERALL,
        "MIN_BINS_SEG": MIN_BINS_SEG,
        "MIN_BINS_HALF": MIN_BINS_HALF,
        "CLASS_ORDER": "WCW -> MHM -> CCC -> WWW -> MIX -> INSUF",
        "NOTE": (
            "INSUF=Cannot compute the slope for overall (1-12k) or for any pattern segment (7-11 / 4-8 / 1-5);"
            "half(1-6/6-12) half-window slopes are used only for WCW constraints and MHM half-window split handling, and are not part of the INSUF criterion."
        ),
    }

    summary_path = os.path.join(out_root, f"{short_base}_SUMMARY_{ts}.xlsx")
    build_summary_workbook(summary_path, params=params, rows=summary_rows, class_counts=class_counts)

    # Console summary
    valid_total = total_blocks - class_counts.get("INSUF", 0)
    print(f"[DONE] Output root: {out_root}")
    print(f"[INFO] total blocks (series) written: {total_blocks}")
    print(f"[INFO] valid_total (non-INSUF): {valid_total} | INSUF: {class_counts.get('INSUF', 0)}")
    print("===== FINAL CLASS COUNTS =====")
    for cls in CLASS_ORDER:
        cnt = class_counts.get(cls, 0)
        if cls != "INSUF" and valid_total > 0:
            print(f"{cls:>6s} : {cnt:4d}  ({cnt/valid_total:.2%} of valid)")
        else:
            print(f"{cls:>6s} : {cnt:4d}")
    print(f"[DONE] Summary saved: {summary_path}")


if __name__ == "__main__":
    main()
