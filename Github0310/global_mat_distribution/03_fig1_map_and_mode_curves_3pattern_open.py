# -*- coding: utf-8 -*-
"""
Fig1 (3Pattern): A (Map) + B (Right stacked curves)  [ABSOLUTE LAYOUT, RECORD COUNT]

3-pattern merge:
- MHM stays
- WARMING = WWW + WCW
- COOLING = CCC + MIX

A legend: n = a+b   (a=dataset records, b=manual-added records incl. SN)
B labels: n = total (a+b)  -> shown as n=total ONLY

Style rules kept:
- Dataset points: circles (color by pattern) with black edge (lw=2)
- Manual points: circles with visible outline (lw=1)
- FIT manual points: triangles (^), red outline (lw=2), s=120
- KC manual point: square (s), red outline (lw=2), s=90
- SN: red star (size=210, yellow outline), counted into WARM manual-added
- EPS output enabled
"""

import os
from pathlib import Path
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

import cartopy.crs as ccrs
import cartopy.feature as cfeature


# ===================== ① ONLY EDIT HERE =====================
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_ROOT_DIR = str(SCRIPT_DIR)
# ============================================================


# ===================== ② ALL NUMBERS TO EDIT (central parameter block) =====================
plt.rcParams["font.family"] = "Arial"
plt.rcParams["axes.unicode_minus"] = False

# output
SAVE_PNG = True
SAVE_EPS = True
DPI = 320
FIGSIZE = (14.2, 6.8)

# absolute layout (A/B same height)
FIG_LEFT, FIG_RIGHT = 0.03, 0.98
FIG_BOTTOM, FIG_TOP = 0.03, 0.99
A_WIDTH = 0.71
AB_GAP = 0.058  # A-B spacing (larger value = greater separation)

# map point style
POINT_SIZE = 40
BASE_ALPHA = 0.95

# Dataset points: edge width = 0
EDGE_LW_BASE = 1.0
EDGE_COLOR_BASE = "black"

# Manual points: make outlines clearly visible (same width)
EDGE_LW_MANUAL = 1
EDGE_COLOR_MANUAL = "white"   # More visible (you may also change it to yellow / black)

# special manual styles
FIT_MARKER = "^"
FIT_TRI_SIZE = 120

KC_SQUARE_SIZE = 90          # Set KC size to 90 as well
SPECIAL_EDGE_COLOR = "yellow"   # Triangle/square edge color
SPECIAL_EDGE_LW = 2       # Line width = 2

# SN star
SN_STAR = dict(lon=117.07, lat=28.72, s=250, marker="*", fc="red", ec="black", lw=1.5, z=20)
SN_TEXT = dict(dx=-10.0, dy=3.0, fontsize=15)

# latitude dashed lines (6 bands boundaries)
LAT_BAND_BOUNDARIES = [-60.0, -30.0, 0.0, 30.0, 60.0]
LATBAND_LW = 0.70
LATBAND_COLOR = "0.35"
LATBAND_ALPHA = 0.55

# panel letters
PANEL_LETTER_POS = {"A": (0.012, 0.988), "B": (-0.20, 0.988)}
PANEL_LETTER_FS = 17

# legend
LEGEND_LOC = "lower left"
LEGEND_FS = 14
LEGEND_MARKERSCALE = 1.4

# curves
BIN_WIDTH_Y = 500.0
CURVE_LO = 0.0
CURVE_HI = 12000.0
MIN_VALID_BINS_FOR_CURVE = 3
CURVE_GREY_ALPHA = 0.55
CURVE_GREY_LW = 0.60
CURVE_MEAN_LW = 1.90
CURVE_GRID_ALPHA = 0.18
CURVE_XLABEL = "Age (ka BP)"
CURVE_YLABEL = "Z-score"

# stacked axes inside B  (✅ 3 patterns now)
STACK_VGAP = 0.012
# ============================================================


# ===================== ③ ORIGINAL CLASSES & 3PATTERN =====================
ORIG_CLASSES = ["MHM", "MIX", "WCW", "WWW", "CCC"]

PATTERNS = {
    "MHM":  ["MHM"],
    "WARMING": ["WWW", "WCW"],
    "COOLING": ["CCC", "MIX"],
}
PATTERN_ORDER = ["MHM", "WARMING", "COOLING"]
STACK_N = len(PATTERN_ORDER)

# Colors follow your request: WCW/WWW = red; CCC/MIX = blue; MHM = dark yellow (clear in PPT)
PATTERN_COLOR = {
    "WARMING": "#B11226",  # deep red
    "COOLING": "#0B3D91",  # deep blue
    "MHM":  "#C9A227",  # deep yellow (low saturation but clear)
}

# special manual names
FIT_NAMES = {"LL", "MC-FIT", "MEN-FIT"}   # FIT triangles
SQUARE_NAMES = {"KC"}                    # KC square


def orig_to_pattern(orig_cls: str) -> str:
    for pat, members in PATTERNS.items():
        if orig_cls in members:
            return pat
    raise ValueError(f"Unknown original class: {orig_cls}")


# ===================== ④ MANUAL POINTS =====================
# format: (name, lat, lon, original_class, countable)
MANUAL_POINTS: List[Tuple[str, float, float, str, int]] = []

# WCW manual points (countable; SN counted but plotted as star only)
for name, lat, lon in [
    ("LL",  -8.5333, 120.4333),
    ("SN",  28.72,   117.07),
    ("SZY", 25.53,   116.68),
    ("NYH", 31.10,   118.95),
    ("GB2-GC1", 26.67, -93.92),
    ("KC", 34.6167, 136.7667),
]:
    MANUAL_POINTS.append((name, lat, lon, "WCW", 1))

# WWW manual points (countable)
for name, lat, lon in [
    ("SCS-T", 20.11667, 117.38333),
    ("ECS-T", 31.64, 128.94),
    ("HKUV11", 22.21, 113.89),
    ("NS02G", 19.8, 113.9),
    ("2904", 19.4553, 116.2525),
    ("2905", 20.136, 117.36),
    ("18252-3", 9.2333, 109.38333),
    ("18287-3", 5.65, 110.65),
    ("10043", -7.3095, 105.0588333),
    ("B-3GC", 31.4895, 128.519),
    ("MD052928", -11.2877, 148.86),
    ("SO18517", -1.536633, 117.5626),
    ("GeoB 10029-4", -1.5, 100.1333333333),
    ("GeoB 10038-4", -5.9333333333, 103.25),
]:
    MANUAL_POINTS.append((name, lat, lon, "WWW", 1))

# FIT emphasis points (MIX)  —— MC-FIT NOT COUNTED
MANUAL_POINTS.append(("MC-FIT", 47.49, 7.02, "MIX", 0))
MANUAL_POINTS.append(("MEN-FIT", 42.967361, -1.884583, "MIX", 1))


# ===================== ⑤ REGEX & HELPERS =====================
LONLAT_RE = re.compile(r"Lon\s*=\s*([-+]?\d*\.?\d+)\s*,\s*Lat\s*=\s*([-+]?\d*\.?\d+)", re.I)


def norm_lon(lon: float) -> float:
    if lon is None or not np.isfinite(lon):
        return lon
    lon = float(lon)
    if lon > 180.0:
        lon -= 360.0
    if lon < -180.0:
        lon += 360.0
    return lon


def parse_lon_lat(coord_str: str) -> Optional[Tuple[float, float]]:
    if not coord_str:
        return None
    m = LONLAT_RE.search(str(coord_str))
    if not m:
        return None
    lon = norm_lon(float(m.group(1)))
    lat = float(m.group(2))
    return lon, lat


def iter_blocks_row4(ws) -> List[int]:
    starts = []
    for c in range(1, ws.max_column):
        v1 = str(ws.cell(4, c).value).strip().lower()
        v2 = str(ws.cell(4, c + 1).value).strip().lower()
        if v1 == "age" and v2 == "value":
            starts.append(c)
    return starts


def find_latest_xlsx(folder: str, class_code: str) -> str:
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"Missing folder: {folder}")

    cands = []
    for fn in os.listdir(folder):
        if not fn.lower().endswith(".xlsx"):
            continue
        low = fn.lower()
        if "summary" in low or "total" in low:
            continue
        if f"_{class_code}_" not in fn:
            continue
        fp = os.path.join(folder, fn)
        if os.path.isfile(fp):
            cands.append(fp)

    if not cands:
        raise FileNotFoundError(f"No xlsx with _{class_code}_ found in {folder}")

    cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return cands[0]


def extract_base_points_from_xlsx(xlsx_path: str, orig_cls: str) -> pd.DataFrame:
    """
    IMPORTANT:
    - Count is RECORD count (series blocks), NOT site count.
    - DO NOT de-dup by lon/lat.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    rows = []
    for sh in wb.sheetnames:
        ws = wb[sh]
        starts = iter_blocks_row4(ws)
        for c in starts:
            coord = ws.cell(2, c).value
            ll = parse_lon_lat(coord)
            if ll is None:
                continue
            lon, lat = ll
            rows.append({"orig_class": orig_cls, "lon": lon, "lat": lat})
    return pd.DataFrame(rows)


def build_manual_df() -> pd.DataFrame:
    rows = []
    for name, lat, lon, orig_cls, countable in MANUAL_POINTS:
        pat = orig_to_pattern(orig_cls)
        rows.append({
            "name": name,
            "orig_class": orig_cls,
            "pattern": pat,
            "lon": norm_lon(float(lon)),
            "lat": float(lat),
            "countable": int(countable),
            "is_fit": int(str(name) in FIT_NAMES),
            "is_square": int(str(name) in SQUARE_NAMES),
        })
    return pd.DataFrame(rows)


def manual_add_count(base_df: pd.DataFrame, manual_df_pat: pd.DataFrame) -> int:
    """
    b = manual-added count (countable==1) and NOT already present in base by rounding.
    """
    if manual_df_pat is None or manual_df_pat.empty:
        return 0
    mdf = manual_df_pat[manual_df_pat["countable"] == 1].copy()
    if mdf.empty:
        return 0

    # de-dup manual itself
    mdf["lon_r"] = mdf["lon"].round(4)
    mdf["lat_r"] = mdf["lat"].round(4)
    mdf = mdf.drop_duplicates(subset=["lon_r", "lat_r"]).reset_index(drop=True)

    if base_df is None or base_df.empty:
        return int(len(mdf))

    b = base_df.copy()
    b["lon_r"] = b["lon"].round(4)
    b["lat_r"] = b["lat"].round(4)
    base_set = set(zip(b["lon_r"].values.tolist(), b["lat_r"].values.tolist()))

    add = 0
    for lonr, latr in zip(mdf["lon_r"].values.tolist(), mdf["lat_r"].values.tolist()):
        if (lonr, latr) not in base_set:
            add += 1
    return int(add)


def add_panel_letter(ax, letter: str):
    x, y = PANEL_LETTER_POS.get(letter, (0.012, 0.988))
    ax.text(x, y, letter, transform=ax.transAxes,
            ha="left", va="top", fontsize=PANEL_LETTER_FS, fontweight="bold")


def draw_latband_dashed(ax):
    lons = np.linspace(-180, 180, 361)
    for lat in LAT_BAND_BOUNDARIES:
        ax.plot(lons, np.full_like(lons, lat),
                transform=ccrs.PlateCarree(),
                linestyle="--", linewidth=LATBAND_LW,
                color=LATBAND_COLOR, alpha=LATBAND_ALPHA, zorder=1)


def draw_manual_lat_labels(ax, lats=(-60, -30, 0, 30, 60), lon=-170, fontsize=9):
    for lat in lats:
        ax.text(lon, lat, f"{lat}°", transform=ccrs.PlateCarree(),
                ha="right", va="center", fontsize=fontsize, color="black", zorder=50)


# ===================== ⑥ CURVE EXTRACTION =====================
def maybe_ka_to_years(arr: np.ndarray) -> np.ndarray:
    if arr.size == 0:
        return arr
    mx = np.nanmax(arr)
    if np.isfinite(mx) and mx <= 200.0:
        return arr * 1000.0
    return arr


def extract_series_block(ws, start_col: int) -> Tuple[List[float], List[float]]:
    ages, vals = [], []
    max_row = ws.max_row

    last = 4
    for r in range(max_row, 4, -1):
        if ws.cell(r, start_col).value is not None or ws.cell(r, start_col + 1).value is not None:
            last = r
            break
    if last <= 4:
        return ages, vals

    for r in range(5, last + 1):
        a = ws.cell(r, start_col).value
        v = ws.cell(r, start_col + 1).value
        if a is None or v is None:
            continue
        try:
            af = float(a)
            vf = float(v)
        except Exception:
            continue
        if not (np.isfinite(af) and np.isfinite(vf)):
            continue
        ages.append(af)
        vals.append(vf)
    return ages, vals


def bin_mean_series(ages: List[float], vals: List[float], lo: float, hi: float, bin_width: float):
    x = np.asarray(ages, dtype=float)
    y = np.asarray(vals, dtype=float)
    x = maybe_ka_to_years(x)

    m = np.isfinite(x) & np.isfinite(y) & (x >= lo) & (x <= hi)
    x = x[m]
    y = y[m]

    edges = np.arange(lo, hi + bin_width, bin_width)
    centers = edges[:-1] + bin_width / 2.0
    means = np.full_like(centers, np.nan, dtype=float)

    if x.size == 0:
        return centers, means

    for i in range(len(centers)):
        a0, a1 = edges[i], edges[i + 1]
        mm = (x >= a0) & (x < a1) if i < len(centers) - 1 else (x >= a0) & (x <= a1)
        if np.any(mm):
            means[i] = float(np.mean(y[mm]))
    return centers, means


def zscore_series(y: np.ndarray) -> np.ndarray:
    yy = np.asarray(y, dtype=float)
    m = np.isfinite(yy)
    if np.sum(m) < MIN_VALID_BINS_FOR_CURVE:
        return np.full_like(yy, np.nan)
    mu = float(np.nanmean(yy))
    sd = float(np.nanstd(yy))
    if (not np.isfinite(sd)) or sd == 0.0:
        return np.full_like(yy, np.nan)
    return (yy - mu) / sd


def extract_all_zcurves_from_xlsx(xlsx_path: str):
    """
    Return:
      x_kyr, curves(list), mean_curve, n_used (dataset curves count)
    """
    wb = load_workbook(xlsx_path, data_only=True)
    centers_ref = None
    curves = []

    for sh in wb.sheetnames:
        ws = wb[sh]
        starts = iter_blocks_row4(ws)
        for c in starts:
            ages, vals = extract_series_block(ws, c)
            if not ages or not vals:
                continue
            centers, means = bin_mean_series(ages, vals, CURVE_LO, CURVE_HI, BIN_WIDTH_Y)
            if centers_ref is None:
                centers_ref = centers.copy()
            else:
                if len(centers) != len(centers_ref):
                    continue
            z = zscore_series(means)
            if np.all(~np.isfinite(z)):
                continue
            curves.append(z)

    if centers_ref is None or len(curves) == 0:
        return None, [], None, 0

    M = np.vstack(curves)
    mean_curve = np.nanmean(M, axis=0)
    x_kyr = centers_ref / 1000.0
    return x_kyr, curves, mean_curve, int(len(curves))


def extract_all_zcurves_from_many_xlsx(xlsx_list: List[str]):
    """
    Merge multiple xlsx into one pattern:
    return x_kyr, all_curves(list), mean_curve, n_used
    """
    x_ref = None
    all_curves = []

    for fp in xlsx_list:
        x, curves, _mean, n_used = extract_all_zcurves_from_xlsx(fp)
        if x is None or len(curves) == 0:
            continue
        if x_ref is None:
            x_ref = x
        else:
            if len(x) != len(x_ref):
                continue
        all_curves.extend(curves)

    if x_ref is None or len(all_curves) == 0:
        return None, [], None, 0

    M = np.vstack(all_curves)
    mean_curve = np.nanmean(M, axis=0)
    return x_ref, all_curves, mean_curve, int(len(all_curves))


# ===================== ⑦ PLOTTING =====================
def plot_map_panel(ax, base_points_pat: Dict[str, pd.DataFrame], manual_df: pd.DataFrame,
                   legend_counts_ab: Dict[str, Tuple[int, int]]):
    ax.set_global()
    ax.add_feature(cfeature.OCEAN, linewidth=0)
    ax.add_feature(cfeature.LAND, linewidth=0)
    ax.add_feature(cfeature.COASTLINE, linewidth=0.6)
    ax.add_feature(cfeature.BORDERS, linewidth=0.35)

    draw_latband_dashed(ax)

    # latitude labels only
    try:
        gl = ax.gridlines(crs=ccrs.PlateCarree(), draw_labels=True, linewidth=0.0, color="none")
        gl.top_labels = False
        gl.right_labels = False
        gl.bottom_labels = False
        gl.left_labels = True
        gl.xlabel_style = {"size": 0}
        gl.ylabel_style = {"size": 15}
    except Exception:
        draw_manual_lat_labels(ax, lats=(-60, -30, 0, 30, 60), lon=-170, fontsize=9)

    handles, labels = [], []

    # --- dataset points (circles) ---
    for pat in PATTERN_ORDER:
        dfb = base_points_pat.get(pat, pd.DataFrame(columns=["lon", "lat"]))
        color = PATTERN_COLOR[pat]

        a, b = legend_counts_ab.get(pat, (0, 0))
        lab = f"{pat} (n={a}+{b})"  # A-panel legend must show a+b

        if not dfb.empty:
            h = ax.scatter(
                dfb["lon"].values, dfb["lat"].values,
                s=POINT_SIZE, marker="o",
                facecolor=color,
                edgecolor=EDGE_COLOR_BASE, linewidth=EDGE_LW_BASE,  # dataset-point edge width = 0
                alpha=BASE_ALPHA,
                transform=ccrs.PlateCarree(),
                zorder=3
            )
        else:
            h = ax.scatter(
                [], [],
                s=POINT_SIZE, marker="o",
                facecolor=color,
                edgecolor=EDGE_COLOR_BASE, linewidth=EDGE_LW_BASE,
                alpha=BASE_ALPHA,
                transform=ccrs.PlateCarree(),
                zorder=3
            )

        handles.append(h)
        labels.append(lab)

    # --- manual overlay ---
    if manual_df is not None and (not manual_df.empty):
        # Do not plot SN as a circle/triangle/square (to avoid overlap with the red star), but it is already counted in b
        mplot = manual_df[manual_df["name"].astype(str) != "SN"].copy()

        for pat in PATTERN_ORDER:
            sub = mplot[mplot["pattern"] == pat]
            if sub.empty:
                continue

            # squares (KC) red edge
            sub_sq = sub[sub["is_square"] == 1]
            if not sub_sq.empty:
                ax.scatter(
                    sub_sq["lon"].values, sub_sq["lat"].values,
                    s=KC_SQUARE_SIZE, marker="s",
                    facecolor=PATTERN_COLOR[pat],
                    edgecolor=SPECIAL_EDGE_COLOR, linewidth=SPECIAL_EDGE_LW,
                    alpha=1.0, transform=ccrs.PlateCarree(),
                    zorder=7, label="_nolegend_"
                )

            # FIT triangles red edge
            sub_fit = sub[sub["is_fit"] == 1]
            if not sub_fit.empty:
                ax.scatter(
                    sub_fit["lon"].values, sub_fit["lat"].values,
                    s=FIT_TRI_SIZE, marker=FIT_MARKER,
                    facecolor=PATTERN_COLOR[pat],
                    edgecolor=SPECIAL_EDGE_COLOR, linewidth=SPECIAL_EDGE_LW,
                    alpha=1.0, transform=ccrs.PlateCarree(),
                    zorder=7, label="_nolegend_"
                )

            # other manual circles visible edge
            sub_circle = sub[(sub["is_square"] == 0) & (sub["is_fit"] == 0)]
            if not sub_circle.empty:
                ax.scatter(
                    sub_circle["lon"].values, sub_circle["lat"].values,
                    s=POINT_SIZE, marker="o",
                    facecolor=PATTERN_COLOR[pat],
                    edgecolor=EDGE_COLOR_MANUAL, linewidth=EDGE_LW_MANUAL,
                    alpha=1.0, transform=ccrs.PlateCarree(),
                    zorder=6, label="_nolegend_"
                )

    # SN star (belongs to WARMING)
    ax.scatter(
        [SN_STAR["lon"]], [SN_STAR["lat"]],
        s=SN_STAR["s"], marker=SN_STAR["marker"],
        facecolor=SN_STAR["fc"], edgecolor=SN_STAR["ec"], linewidth=SN_STAR["lw"],
        transform=ccrs.PlateCarree(), zorder=SN_STAR["z"], label="_nolegend_"
    )
    ax.text(
        SN_STAR["lon"] + SN_TEXT["dx"], SN_STAR["lat"] + SN_TEXT["dy"], "SN",
        transform=ccrs.PlateCarree(), fontsize=SN_TEXT["fontsize"],
        fontweight="bold", color="black", zorder=SN_STAR["z"] + 1
    )

    ax.legend(
        handles, labels,
        loc="lower left",
        bbox_to_anchor=(0.01, 0.0),  # slight position adjustment
        fontsize=LEGEND_FS,
        frameon=True,
        markerscale=LEGEND_MARKERSCALE
    )


def plot_one_curve_axis(ax, pat: str, curve_pack: Dict[str, dict], total_n: int):
    x = curve_pack[pat]["x"]
    curves = curve_pack[pat]["curves"]
    mean_curve = curve_pack[pat]["mean"]

    if x is not None and len(curves) > 0:
        for z in curves:
            ax.plot(x, z, color="0.75", linewidth=CURVE_GREY_LW, alpha=CURVE_GREY_ALPHA, zorder=1)
        ax.plot(x, mean_curve, color="black", linewidth=CURVE_MEAN_LW, zorder=3)

    # ax.invert_xaxis()  # commented out: do not invert the X axis
    ax.set_xlim(0, CURVE_HI / 1000.0)  # ✅ 0 -> 12k（ka BP）
    ax.grid(True, alpha=CURVE_GRID_ALPHA)
    ax.tick_params(axis="both", labelsize=15)  # change 9 to any font size you prefer, such as 10/11
    # Panel B shows only the total
    ax.text(
        0.01, 0.75, f"{pat} (n={total_n})",
        transform=ax.transAxes, fontsize=14, fontweight="bold",
        color=PATTERN_COLOR[pat]
    )

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)


def create_right_stacked_axes(fig, rect: Tuple[float, float, float, float]) -> List[plt.Axes]:
    x0, y0, w, h = rect
    n = STACK_N
    gap = STACK_VGAP
    each_h = (h - gap * (n - 1)) / n

    axs = []
    for i in range(n):
        y = y0 + h - (i + 1) * each_h - i * gap
        axs.append(fig.add_axes([x0, y, w, each_h]))
    return axs


# ===================== MAIN =====================
def main():
    if not os.path.isdir(OUT_ROOT_DIR):
        raise FileNotFoundError(f"OUT_ROOT_DIR not found: {OUT_ROOT_DIR}")

    # ---- output folder: YYYYMMDD_HHMM_3Pattern ----
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    OUT_DIR = os.path.join(OUT_ROOT_DIR, f"{ts}_3Pattern")
    os.makedirs(OUT_DIR, exist_ok=True)

    # locate latest xlsx for original classes
    orig_xlsx: Dict[str, str] = {}
    orig_base_points: Dict[str, pd.DataFrame] = {}

    for oc in ORIG_CLASSES:
        folder = os.path.join(OUT_ROOT_DIR, oc)
        fp = find_latest_xlsx(folder, oc)
        orig_xlsx[oc] = fp
        orig_base_points[oc] = extract_base_points_from_xlsx(fp, oc)

    manual_df = build_manual_df()

    # ---- merge base points to 3 patterns (record count, no de-dup) ----
    base_points_pat: Dict[str, pd.DataFrame] = {}
    pattern_xlsx_list: Dict[str, List[str]] = {}

    for pat in PATTERN_ORDER:
        members = PATTERNS[pat]
        dfs = [orig_base_points[m] for m in members if (m in orig_base_points and orig_base_points[m] is not None)]
        base_points_pat[pat] = pd.concat(dfs, ignore_index=True) if len(dfs) > 0 else pd.DataFrame(columns=["lon", "lat"])
        pattern_xlsx_list[pat] = [orig_xlsx[m] for m in members]

    # ---- A legend counts: n=a+b ----
    legend_counts_ab: Dict[str, Tuple[int, int]] = {}
    manual_added_by_pat: Dict[str, int] = {}

    for pat in PATTERN_ORDER:
        dfb = base_points_pat.get(pat, pd.DataFrame())
        a = int(len(dfb)) if (dfb is not None and not dfb.empty) else 0

        msub = manual_df[manual_df["pattern"] == pat] if (manual_df is not None and not manual_df.empty) else pd.DataFrame()
        b = manual_add_count(dfb, msub)

        legend_counts_ab[pat] = (a, b)
        manual_added_by_pat[pat] = b

    # ---- curves: merge for each pattern ----
    curve_pack: Dict[str, dict] = {}
    dataset_curve_n: Dict[str, int] = {}

    for pat in PATTERN_ORDER:
        x, curves, mean_curve, n_used = extract_all_zcurves_from_many_xlsx(pattern_xlsx_list[pat])
        curve_pack[pat] = {"x": x, "curves": curves, "mean": mean_curve}
        dataset_curve_n[pat] = int(n_used)

    # ---- B total n = dataset curves + manual-added (MC-FIT excluded by countable=0) ----
    total_curve_n: Dict[str, int] = {}
    for pat in PATTERN_ORDER:
        total_curve_n[pat] = dataset_curve_n[pat] + manual_added_by_pat.get(pat, 0)

    # output files
    out_png = os.path.join(OUT_DIR, f"Fig1_AB_MapPlusCurves_{ts}_3Pattern.png")
    out_eps = os.path.join(OUT_DIR, f"Fig1_AB_MapPlusCurves_{ts}_3Pattern.eps")

    # ---- absolute layout ----
    fig = plt.figure(figsize=FIGSIZE, dpi=DPI)
    fig.patch.set_facecolor("white")

    ax_h = FIG_TOP - FIG_BOTTOM
    A_rect = (FIG_LEFT, FIG_BOTTOM, A_WIDTH, ax_h)

    B_x0 = FIG_LEFT + A_WIDTH + AB_GAP
    B_w = FIG_RIGHT - B_x0
    if B_w <= 0.05:
        raise RuntimeError("B panel width too small. Reduce A_WIDTH or AB_GAP.")
    # --- Panel B height control (A full height, B shorter) ---
    B_H_SCALE = 0.80  # Panel-B height scale (0~1); 0.80~0.90 is usually comfortable
    B_h = ax_h * B_H_SCALE
    B_y0 = FIG_BOTTOM + (ax_h - B_h) / 2.0  # vertically centered
    B_rect = (B_x0, B_y0, B_w, B_h)

    # panel A
    axA = fig.add_axes(list(A_rect), projection=ccrs.Robinson())
    plot_map_panel(axA, base_points_pat=base_points_pat, manual_df=manual_df, legend_counts_ab=legend_counts_ab)
    add_panel_letter(axA, "A")

    # panel B (3 stacked)
    axsB = create_right_stacked_axes(fig, B_rect)
    for i, pat in enumerate(PATTERN_ORDER):
        plot_one_curve_axis(axsB[i], pat, curve_pack=curve_pack, total_n=total_curve_n[pat])

        if i < len(PATTERN_ORDER) - 1:
            axsB[i].set_xticklabels([])
        else:
            axsB[i].set_xlabel(CURVE_XLABEL, fontsize=15)

        if i == 1:
            axsB[i].set_ylabel(CURVE_YLABEL, fontsize=15)
        else:
            axsB[i].set_ylabel("")

    add_panel_letter(axsB[0], "B")

    # save
    if SAVE_PNG:
        fig.savefig(out_png, dpi=DPI, bbox_inches="tight")
        print("[DONE] PNG:", out_png)
    if SAVE_EPS:
        fig.savefig(out_eps, format="eps", bbox_inches="tight")
        print("[DONE] EPS:", out_eps)

    plt.close(fig)

    print("\n[CHECK] A legend counts (n=a+b):")
    for pat in PATTERN_ORDER:
        a, b = legend_counts_ab[pat]
        print(f"  {pat}: {a}+{b} = {a+b}")

    print("\n[CHECK] B curve labels (n=total):")
    for pat in PATTERN_ORDER:
        print(f"  {pat}: total={total_curve_n[pat]}  (dataset={dataset_curve_n[pat]} + manual={manual_added_by_pat[pat]})")

    print("\n[INFO] Output folder:", OUT_DIR)


if __name__ == "__main__":
    main()
