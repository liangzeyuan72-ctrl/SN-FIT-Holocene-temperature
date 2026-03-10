# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from pathlib import Path
import json
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# ===================== Edit here only =====================
SCRIPT_DIR = Path(__file__).resolve().parent
LIPD_DIR = str(SCRIPT_DIR)
METADATA_XLSX = str(SCRIPT_DIR / "Temp12k_v1_essential_metadata_NOAA.xlsx")  # Place the metadata file in the script directory
# ==========================================================

# ===================== Annual filter switch =====================
FILTER_ANNUAL_ONLY = True   # True: keep only Season General=annual; False: do not filter
ANNUAL_KEYWORDS = ["annual"]
# ==========================================================

# ===================== Composite priority rule =====================
PREFER_TEMPERATURE_COMPOSITE = True   # True: if temperatureComposite exists for the same site, keep only that record
# ============================================================

# ===================== Added: 0–12k coverage filter (1k bins) =====================
FILTER_0_12K_KYR_COVERAGE = True

# 0–12k (year BP) uses 1000-year bins, for a total of 12 bins: [0-1k), [1-2k), ..., [11-12k]
COVER_LO_YBP = 0.0
COVER_HI_YBP = 12000.0
BIN_WIDTH_Y = 1000.0

# Rule: each non-empty bin must contain at least 2 points; at most 4 bins may be empty; and 3 consecutive empty bins are not allowed (max consecutive empty bins <= 2)
MIN_PTS_PER_NONEMPTY_BIN = 2
MAX_EMPTY_BINS = 4
MAX_CONSEC_EMPTY_BINS = 2
# ============================================================================

# Optional: write unit/filter log for checking
WRITE_UNIT_LOG = True
UNIT_LOG_CSV = "unit_alignment_log.csv"

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLS = 16_384
BAD_SHEET_CHARS = [":", "\\", "/", "?", "*", "[", "]"]

# degC / Kelvin / Fahrenheit detection
DEGC_KEYS = ["degc", "°c", "celsius", "degree c", "degrees c", "deg c", "degree_celsius", "degrees_celsius"]
KELVIN_KEYS = ["kelvin", "k"]
FAHRENHEIT_KEYS = ["fahrenheit", "°f", "degf", "deg f", "degree f", "degrees f"]

# Temperature-column keywords (fallback when units are unreliable)
TEMP_KEYS = ["temperature", "temp", "sst", "sat", "gmst", "tmean", "tas"]

# Strong exclusion keywords explicitly indicating non-temperature variables (to avoid misidentifying δ18O/isotopes, etc.)
NON_TEMP_HINT_KEYS = [
    "d18o", "δ18", "d2h", "δd", "permil", "‰", "vpdb", "vsmow",
    "isotope", "isotopic", "oxygen isotope", "carbon isotope",
    "ratio"
]

# Error/uncertainty-column keywords (strong exclusion)
UNCERTAINTY_KEYS = [
    "uncert", "uncertainty",
    "error", "err",
    "sigma", "1sigma", "2sigma",
    "sd", "std", "stdev", "stderr", "se",
    "rmse",
    "confidence", "ci", "95%",
    "upper", "lower", "bounds", "bound",
    "variance", "var",
    "precision"
]

# Anomaly keywords (to avoid subtracting 273.15 from Kelvin anomalies)
ANOM_KEYS = ["anom", "anomaly", "departure", "relative", "offset"]

# Time-axis priority (LiPD columns.variableName)
TIME_PRIORITY = ["ageDuplicate", "ageMedian", "age", "ageOriginal", "year", "time", "yr", "t"]


def _norm_key(s: str) -> str:
    """Normalize names by keeping only alphanumeric characters, reducing matching failures caused by dataset/site naming differences."""
    return "".join(ch.lower() for ch in (s or "") if ch.isalnum())


def _norm_pg(pg: str) -> str:
    return (pg or "").strip().lower()


# -------------------- Read Excel (essential metadata) --------------------
def load_essential_metadata(path: str) -> pd.DataFrame:
    """
    Read Temp12k_v1_essential_metadata_NOAA.xlsx and return a DataFrame with flattened column names,
    containing at least: Data Set Name / Site Name / Proxy General / Season General
    """
    raw = pd.read_excel(path, header=None)
    header_row = None
    for i in range(min(30, len(raw))):
        row = raw.iloc[i].astype(str).tolist()
        if any(str(x).strip() == "Data Set Name" for x in row):
            header_row = i
            break
    if header_row is None or header_row < 1:
        df = pd.read_excel(path, header=[3, 4])
    else:
        df = pd.read_excel(path, header=[header_row - 1, header_row])

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [c[1] for c in df.columns]
    else:
        df.columns = [str(c) for c in df.columns]

    if "Data Set Name" not in df.columns:
        raise ValueError("Column 'Data Set Name' was not found in the metadata table. Please check the file format.")

    df = df.dropna(subset=["Data Set Name"])

    for col in ["Data Set Name", "Site Name", "Proxy General"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
        else:
            raise ValueError(f"Missing column in metadata table: {col}")

    # Season General: tolerant matching when possible
    season_col = None
    for c in df.columns:
        if str(c).strip().lower() == "season general":
            season_col = c
            break
    if season_col is None:
        for c in df.columns:
            cc = str(c).strip().lower()
            if ("season" in cc) and ("general" in cc):
                season_col = c
                break

    if season_col is None:
        df["Season General"] = ""
    else:
        df["Season General"] = df[season_col].astype(str).str.strip()

    return df


def build_proxy_general_lookup(meta: pd.DataFrame) -> Tuple[List[str], Dict[str, List[str]], Dict[str, List[str]]]:
    cats = sorted(meta["Proxy General"].dropna().astype(str).str.strip().unique().tolist())

    by_dataset: Dict[str, List[str]] = {}
    by_site: Dict[str, List[str]] = {}

    for _, r in meta.iterrows():
        ds = str(r["Data Set Name"]).strip()
        st = str(r["Site Name"]).strip()
        pg = str(r["Proxy General"]).strip()

        if ds:
            k = ds.lower()
            by_dataset.setdefault(k, [])
            if pg not in by_dataset[k]:
                by_dataset[k].append(pg)

        if st:
            k = st.lower()
            by_site.setdefault(k, [])
            if pg not in by_site[k]:
                by_site[k].append(pg)

    return cats, by_dataset, by_site


# -------------------- Annual filtering: build lookup tables from metadata --------------------
def _is_annual_text(s: str) -> bool:
    txt = (s or "").strip().lower()
    if not txt:
        return False
    return any(k in txt for k in ANNUAL_KEYWORDS)


def build_annual_lookup(
    meta: pd.DataFrame,
) -> Tuple[Dict[str, bool], Dict[str, bool], Dict[Tuple[str, str], bool], Dict[Tuple[str, str], bool]]:
    """
    Return:
      annual_by_dataset: norm(dataset_name) -> True/False
      annual_by_site:    norm(site_name) -> True/False
      annual_by_dataset_pg: (norm(dataset_name), norm(proxy_general)) -> True/False
      annual_by_site_pg:    (norm(site_name),    norm(proxy_general)) -> True/False
    """
    annual_by_dataset: Dict[str, bool] = {}
    annual_by_site: Dict[str, bool] = {}
    annual_by_dataset_pg: Dict[Tuple[str, str], bool] = {}
    annual_by_site_pg: Dict[Tuple[str, str], bool] = {}

    for _, r in meta.iterrows():
        ds = str(r.get("Data Set Name", "")).strip()
        st = str(r.get("Site Name", "")).strip()
        pg = str(r.get("Proxy General", "")).strip()
        sg = str(r.get("Season General", "")).strip()

        is_ann = _is_annual_text(sg)
        ds_k = _norm_key(ds) if ds else ""
        st_k = _norm_key(st) if st else ""
        pg_k = _norm_pg(pg) if pg else ""

        if ds_k:
            annual_by_dataset[ds_k] = annual_by_dataset.get(ds_k, False) or is_ann
            if pg_k:
                key = (ds_k, pg_k)
                annual_by_dataset_pg[key] = annual_by_dataset_pg.get(key, False) or is_ann

        if st_k:
            annual_by_site[st_k] = annual_by_site.get(st_k, False) or is_ann
            if pg_k:
                key = (st_k, pg_k)
                annual_by_site_pg[key] = annual_by_site_pg.get(key, False) or is_ann

    return annual_by_dataset, annual_by_site, annual_by_dataset_pg, annual_by_site_pg


def is_annual_any(
    dataset_name: str,
    site_name: str,
    annual_by_dataset: Dict[str, bool],
    annual_by_site: Dict[str, bool],
) -> bool:
    if dataset_name and annual_by_dataset.get(_norm_key(dataset_name), False):
        return True
    if site_name and annual_by_site.get(_norm_key(site_name), False):
        return True
    return False


def is_annual_for_pg(
    dataset_name: str,
    site_name: str,
    pg: str,
    annual_by_dataset_pg: Dict[Tuple[str, str], bool],
    annual_by_site_pg: Dict[Tuple[str, str], bool],
) -> bool:
    pg_k = _norm_pg(pg)
    if dataset_name and annual_by_dataset_pg.get((_norm_key(dataset_name), pg_k), False):
        return True
    if site_name and annual_by_site_pg.get((_norm_key(site_name), pg_k), False):
        return True
    return False


# -------------------- Read LiPD (measurementTable) --------------------
def _find_member(zf: zipfile.ZipFile, target_basename: str) -> Optional[str]:
    if target_basename in zf.namelist():
        return target_basename
    for nm in zf.namelist():
        if nm.endswith("/" + target_basename) or nm.endswith("\\" + target_basename) or nm.endswith(target_basename):
            return nm
    return None


def _load_metadata_jsonld(zf: zipfile.ZipFile) -> Dict[str, Any]:
    mem = _find_member(zf, "metadata.jsonld")
    if mem is None:
        cands = [n for n in zf.namelist() if n.lower().endswith(".jsonld")]
        if not cands:
            raise ValueError("No metadata.jsonld/.jsonld found in LiPD.")
        mem = cands[0]
    raw = zf.read(mem).decode("utf-8", errors="ignore")
    obj = json.loads(raw)
    if isinstance(obj, list) and len(obj) == 1 and isinstance(obj[0], dict):
        obj = obj[0]
    if not isinstance(obj, dict):
        raise ValueError("metadata jsonld root is not dict.")
    return obj


def _build_graph_index(meta: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    g = meta.get("@graph", None)
    if not isinstance(g, list):
        return {}
    idx = {}
    for node in g:
        if isinstance(node, dict):
            _id = node.get("@id", None)
            if isinstance(_id, str) and _id:
                idx[_id] = node
    return idx


def _resolve_ref(obj: Any, graph_idx: Dict[str, Dict[str, Any]]) -> Any:
    if isinstance(obj, dict) and set(obj.keys()) == {"@id"}:
        _id = obj.get("@id")
        if isinstance(_id, str) and _id in graph_idx:
            return graph_idx[_id]
    return obj


def _find_dataset_node(meta: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(meta.get("@graph", None), list):
        return meta

    best = None
    best_score = -1
    for node in meta["@graph"]:
        if not isinstance(node, dict):
            continue
        score = 0
        for k in ["dataSetName", "paleoData", "chronData", "geo", "site", "archiveType"]:
            if k in node:
                score += 1
        t = node.get("@type", "")
        if isinstance(t, str) and "Dataset" in t:
            score += 1
        if score > best_score:
            best, best_score = node, score
    return best if isinstance(best, dict) else meta


def _iter_measurement_tables(dataset: Dict[str, Any], graph_idx: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    tables: List[Dict[str, Any]] = []

    paleo = dataset.get("paleoData", None)
    if isinstance(paleo, dict):
        paleo = [paleo]
    if not isinstance(paleo, list):
        paleo = []

    def grab(container: Dict[str, Any]):
        for key in ["measurementTable", "measurementTables", "paleoDataTable", "paleoDataTables"]:
            v = container.get(key, None)
            if v is None:
                continue
            if isinstance(v, dict):
                v = [_resolve_ref(v, graph_idx)]
            elif isinstance(v, list):
                v = [_resolve_ref(x, graph_idx) for x in v]
            else:
                continue
            for t in v:
                if isinstance(t, dict):
                    tables.append(t)

    for item in paleo:
        if isinstance(item, dict):
            item = _resolve_ref(item, graph_idx)
            if isinstance(item, dict):
                grab(item)

    if not tables:
        grab(dataset)

    return tables


def _table_filename(t: Dict[str, Any]) -> Optional[str]:
    for k in ["filename", "fileName", "dataFile", "datafile", "file", "path"]:
        v = t.get(k, None)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _table_missing_value(t: Dict[str, Any]) -> Optional[Any]:
    for k in ["missingValue", "missing_value", "NAValue", "na_value"]:
        if k in t:
            return t.get(k)
    return None


def _table_columns(t: Dict[str, Any]) -> List[Dict[str, Any]]:
    for k in ["columns", "paleoDataColumn", "paleoDataColumns"]:
        v = t.get(k, None)
        if isinstance(v, dict):
            return [v]
        if isinstance(v, list):
            return [c for c in v if isinstance(c, dict)]
    return []


def _col_number(col: Dict[str, Any]) -> Optional[int]:
    for k in ["number", "columnNumber", "colNumber", "column", "idx", "index"]:
        v = col.get(k, None)
        if v is None:
            continue
        try:
            return int(v)
        except Exception:
            pass
    return None


def _col_name(col: Dict[str, Any]) -> str:
    for k in ["variableName", "varName", "name", "variable"]:
        v = col.get(k, None)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _col_units(col: Dict[str, Any]) -> str:
    for k in ["units", "unit", "variableUnits", "variableUnit", "unitName", "unitsName"]:
        v = col.get(k, None)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _col_desc(col: Dict[str, Any]) -> str:
    for k in ["description", "desc", "notes", "note"]:
        v = col.get(k, None)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _col_proxy(col: Dict[str, Any]) -> str:
    for k in ["proxy", "proxyGeneral", "proxyType", "proxyName"]:
        v = col.get(k, None)
        if isinstance(v, str) and v.strip():
            return v.strip()
    v = col.get("proxy", None)
    if isinstance(v, dict):
        for kk in ["proxyGeneral", "proxyType", "proxyName", "name"]:
            vv = v.get(kk, None)
            if isinstance(vv, str) and vv.strip():
                return vv.strip()
        return "proxy(dict)"
    if isinstance(v, list):
        parts = []
        for it in v:
            if isinstance(it, str) and it.strip():
                parts.append(it.strip())
            elif isinstance(it, dict):
                for kk in ["proxyGeneral", "proxyType", "proxyName", "name"]:
                    vv = it.get(kk, None)
                    if isinstance(vv, str) and vv.strip():
                        parts.append(vv.strip())
                        break
        return "; ".join(parts)
    return ""


def _is_degC(units: str) -> bool:
    u = (units or "").strip().lower()
    return bool(u) and any(k in u for k in DEGC_KEYS)


def _is_kelvin(units: str) -> bool:
    u = (units or "").strip().lower()
    return bool(u) and (u == "k" or any(k in u for k in KELVIN_KEYS))


def _is_fahrenheit(units: str) -> bool:
    u = (units or "").strip().lower()
    return bool(u) and any(k in u for k in FAHRENHEIT_KEYS)


def _looks_like_anomaly(text: str) -> bool:
    t = (text or "").lower()
    return any(k in t for k in ANOM_KEYS)


def _looks_like_uncertainty(text: str) -> bool:
    t = (text or "").lower()
    return any(k in t for k in UNCERTAINTY_KEYS)


def _looks_like_non_temp(text: str) -> bool:
    t = (text or "").lower()
    return any(k in t for k in NON_TEMP_HINT_KEYS)


# -------------------- Added: 0–12k coverage filter function --------------------
def _maybe_ka_to_years(ages: np.ndarray) -> np.ndarray:
    """
    Automatically determine whether age is in ka: if max <= 200, treat it as ka and convert to year BP
    """
    if ages.size == 0:
        return ages
    mx = np.nanmax(ages)
    if np.isfinite(mx) and mx <= 200.0:
        return ages * 1000.0
    return ages


def check_kyr_coverage_0_12k(
    ages_raw: np.ndarray,
    lo: float = COVER_LO_YBP,
    hi: float = COVER_HI_YBP,
    bin_width: float = BIN_WIDTH_Y,
    min_pts_nonempty: int = MIN_PTS_PER_NONEMPTY_BIN,
    max_empty_bins: int = MAX_EMPTY_BINS,
    max_consec_empty: int = MAX_CONSEC_EMPTY_BINS,
) -> Tuple[bool, Dict[str, Any]]:
    """
    0–12k (year BP) coverage rules:
    - 12 bins of 1k each
    - each non-empty bin has at least 2 points
    - total number of empty bins <= 4
    - maximum number of consecutive empty bins <= 2 (no 3 consecutive 1-kyr bins without data)
    """
    ages = np.asarray(ages_raw, dtype=float)
    ages = ages[np.isfinite(ages)]
    ages = _maybe_ka_to_years(ages)

    m = (ages >= lo) & (ages <= hi)
    a = ages[m]

    nbins = int((hi - lo) // bin_width)  # 12
    counts = np.zeros(nbins, dtype=int)

    if a.size == 0:
        return False, {
            "counts": counts.tolist(),
            "empty_bins": nbins,
            "max_consecutive_empty": nbins,
            "n_in_window": 0,
            "reason": "no_points_in_0_12k",
        }

    idx = np.floor((a - lo) / bin_width).astype(int)
    idx = np.clip(idx, 0, nbins - 1)
    for k in idx:
        counts[k] += 1

    empty_mask = (counts == 0)
    empty_bins = int(empty_mask.sum())

    bad_nonempty = np.where((counts > 0) & (counts < min_pts_nonempty))[0]
    if bad_nonempty.size > 0:
        return False, {
            "counts": counts.tolist(),
            "empty_bins": empty_bins,
            "max_consecutive_empty": None,
            "n_in_window": int(a.size),
            "reason": f"nonempty_bin_lt_{min_pts_nonempty}",
            "bad_bins": bad_nonempty.tolist(),
        }

    # Count consecutive empty bins
    max_run = 0
    run = 0
    for e in empty_mask:
        run = run + 1 if e else 0
        max_run = max(max_run, run)

    if empty_bins > max_empty_bins:
        return False, {
            "counts": counts.tolist(),
            "empty_bins": empty_bins,
            "max_consecutive_empty": int(max_run),
            "n_in_window": int(a.size),
            "reason": f"empty_bins_gt_{max_empty_bins}",
        }

    if max_run > max_consec_empty:
        return False, {
            "counts": counts.tolist(),
            "empty_bins": empty_bins,
            "max_consecutive_empty": int(max_run),
            "n_in_window": int(a.size),
            "reason": f"consecutive_empty_gt_{max_consec_empty}",
        }

    return True, {
        "counts": counts.tolist(),
        "empty_bins": empty_bins,
        "max_consecutive_empty": int(max_run),
        "n_in_window": int(a.size),
        "reason": "ok",
    }
# ------------------------------------------------------------------


def normalize_temperature_to_degC(values: np.ndarray, units: str, blob_text: str) -> Tuple[np.ndarray, str, str]:
    """
    Align temperature to degC as accurately as possible, and return:
    - new_values (degC or degC_anom)
    - unit_show  ('degC' / 'degC_anom')
    - note       (conversion note)
    """
    u_raw = (units or "").strip()
    blob = (blob_text or "").lower()
    is_anom = _looks_like_anomaly((u_raw + " " + blob).lower())

    if _is_degC(u_raw):
        return values.astype(float), ("degC_anom" if is_anom else "degC"), "units=degC"

    if _is_kelvin(u_raw):
        if is_anom:
            return values.astype(float), "degC_anom", "units=K_anom -> keep as degC_anom"
        return (values.astype(float) - 273.15), "degC", "units=K -> degC (-273.15)"

    if _is_fahrenheit(u_raw):
        if is_anom:
            return (values.astype(float) * (5.0 / 9.0)), "degC_anom", "units=F_anom -> degC_anom (*5/9)"
        return ((values.astype(float) - 32.0) * (5.0 / 9.0)), "degC", "units=F -> degC"

    vv = values.astype(float)
    finite = vv[np.isfinite(vv)]
    if finite.size < 5:
        return vv, ("degC_anom" if is_anom else "degC"), "units=unknown, too_few_points -> assume degC"

    p5, p95 = np.percentile(finite, [5, 95])

    if (p5 > 150) and (p95 < 400):
        if is_anom:
            return vv, "degC_anom", f"units=missing but looks like K_anom (p5={p5:.1f}, p95={p95:.1f})"
        return (vv - 273.15), "degC", f"units=missing, looks like K (p5={p5:.1f}, p95={p95:.1f}) -> -273.15"

    if (p95 > 70) and (p95 < 140) and (p5 > -50):
        return ((vv - 32.0) * (5.0 / 9.0)), "degC", "units=missing, looks like F -> degC"

    return vv, ("degC_anom" if is_anom else "degC"), "units=unknown -> assume degC"


def _read_table_from_zip(zf: zipfile.ZipFile, filename: str, missing_value: Optional[Any]) -> pd.DataFrame:
    mem = _find_member(zf, os.path.basename(filename)) or _find_member(zf, filename)
    if mem is None:
        raise FileNotFoundError(f"Table file not found in zip: {filename}")

    na_values = None
    if missing_value is not None and str(missing_value).strip() != "":
        na_values = [missing_value]

    with zf.open(mem) as f:
        try:
            df = pd.read_csv(
                f, sep=None, engine="python", header=None,
                na_values=na_values, comment="#", on_bad_lines="skip"
            )
        except TypeError:
            f.seek(0)
            df = pd.read_csv(f, sep=None, engine="python", header=None, na_values=na_values, comment="#")
        except Exception:
            f.seek(0)
            try:
                df = pd.read_csv(f, sep=",", engine="python", header=None, na_values=na_values, comment="#")
            except Exception:
                f.seek(0)
                df = pd.read_csv(f, sep="\t", engine="python", header=None, na_values=na_values, comment="#")
    return df


def _extract_lon_lat(dataset: Dict[str, Any], graph_idx: Dict[str, Dict[str, Any]]) -> Tuple[str, str]:
    lon, lat = "", ""
    geo = _resolve_ref(dataset.get("geo", {}), graph_idx)
    if isinstance(geo, dict):
        geom = _resolve_ref(geo.get("geometry", {}), graph_idx)
        if isinstance(geom, dict):
            coords = geom.get("coordinates", None)
            if isinstance(coords, (list, tuple)) and len(coords) >= 2:
                try:
                    lon = f"{float(coords[0]):.5f}"
                    lat = f"{float(coords[1]):.5f}"
                    return lon, lat
                except Exception:
                    pass
        if "longitude" in geo or "latitude" in geo:
            try:
                lon = f"{float(geo.get('longitude')):.5f}"
                lat = f"{float(geo.get('latitude')):.5f}"
                return lon, lat
            except Exception:
                pass
    return lon, lat


def _extract_dataset_and_site_name(dataset: Dict[str, Any], graph_idx: Dict[str, Dict[str, Any]]) -> Tuple[str, str]:
    ds = str(dataset.get("dataSetName", "")).strip()
    site = ""

    geo = _resolve_ref(dataset.get("geo", {}), graph_idx)
    if isinstance(geo, dict):
        site = str(geo.get("siteName", "")).strip()

    if not site:
        site_obj = _resolve_ref(dataset.get("site", {}), graph_idx)
        if isinstance(site_obj, dict):
            site = str(site_obj.get("siteName", "")).strip()

    return ds, site


def pick_time_col(cols_meta: List[Dict[str, Any]]) -> Optional[int]:
    name_to_idx: Dict[str, int] = {}
    for col in cols_meta:
        num = _col_number(col)
        if num is None:
            continue
        name = _col_name(col).strip()
        if name:
            name_to_idx[name.lower()] = num - 1

    for key in TIME_PRIORITY:
        if key.lower() in name_to_idx:
            return name_to_idx[key.lower()]
    return None


def find_temperature_columns(cols_meta: List[Dict[str, Any]]) -> List[Tuple[int, str, str, Dict[str, Any]]]:
    """
    Return:[(idx0, units, blob, col_meta), ...]
    Logic:
    - strongly exclude uncertainty/error columns
    - strongly exclude isotopes / ‰ / VPDB, etc.
    - temperature detection: prioritize units; otherwise use temp keywords as fallback
    """
    out = []
    for col in cols_meta:
        num = _col_number(col)
        if num is None:
            continue
        idx0 = num - 1
        name = _col_name(col)
        units = _col_units(col)
        desc = _col_desc(col)
        proxy = _col_proxy(col)

        blob = f"{name} {desc} {proxy}".strip()
        blob_low = blob.lower()
        units_low = (units or "").lower()
        name_low = (name or "").lower()

        # Exclude error/uncertainty columns
        if _looks_like_uncertainty(blob_low) or _looks_like_uncertainty(units_low) or _looks_like_uncertainty(name_low):
            continue

        # Exclude clearly non-temperature variables (isotopes, etc.)
        if _looks_like_non_temp(blob_low) or _looks_like_non_temp(name_low) or any(k in units_low for k in ["permil", "vpdb", "vsmow", "‰"]):
            continue

        is_temp = _is_degC(units) or _is_kelvin(units) or _is_fahrenheit(units) or any(k in blob_low for k in TEMP_KEYS) or any(k in name_low for k in TEMP_KEYS)
        if not is_temp:
            continue

        out.append((idx0, units, blob, col))
    return out


# -------------------- temperatureComposite detection --------------------
def is_temperature_composite(variable_name: str, blob: str) -> bool:
    vn = (variable_name or "").strip().lower()
    bb = (blob or "").lower()
    return ("temperaturecomposite" in vn) or ("temperaturecomposite" in bb)


# -------------------- Proxy General assignment logic --------------------
def infer_proxy_general_from_blob(blob: str, available: List[str]) -> str:
    b = (blob or "").lower()

    if ("mg/ca" in b) or ("mgca" in b):
        return "Mg/Ca" if "Mg/Ca" in available else available[0]
    if ("alkenone" in b) or ("uk37" in b) or ("uk'37" in b):
        return "alkenone" if "alkenone" in available else available[0]
    if ("tex86" in b) or ("gdgt" in b) or ("brgdt" in b) or ("gddgt" in b):
        return "other biomarker" if "other biomarker" in available else available[0]
    if ("pollen" in b):
        return "pollen" if "pollen" in available else available[0]
    if ("chironom" in b):
        return "chironomid" if "chironomid" in available else available[0]
    if ("ice" in b) or ("greenland" in b) or ("antarctic" in b):
        return "other ice" if "other ice" in available else available[0]
    if ("borehole" in b) or ("tree ring" in b) or ("biophysical" in b):
        return "biophysical" if "biophysical" in available else available[0]

    if "other microfossil" in available:
        return "other microfossil"
    return available[0]


def choose_proxy_general(
    dataset_name: str,
    site_name: str,
    blob: str,
    by_dataset: Dict[str, List[str]],
    by_site: Dict[str, List[str]],
    available: List[str],
) -> str:
    cand: List[str] = []
    if dataset_name:
        cand.extend(by_dataset.get(dataset_name.lower(), []))
    if site_name:
        for x in by_site.get(site_name.lower(), []):
            if x not in cand:
                cand.append(x)

    if not cand:
        return infer_proxy_general_from_blob(blob, available)

    if len(cand) == 1:
        return cand[0] if cand[0] in available else infer_proxy_general_from_blob(blob, available)

    guess = infer_proxy_general_from_blob(blob, available)
    if guess in cand:
        return guess

    for x in cand:
        if x in available:
            return x
    return infer_proxy_general_from_blob(blob, available)


# -------------------- Output Excel: arrange horizontally by sheet --------------------
def _safe_sheet_name(name: str) -> str:
    s = name
    for ch in BAD_SHEET_CHARS:
        s = s.replace(ch, "_")
    s = s.strip() or "sheet"
    return s[:31]


def write_series_block(
    ws,
    start_col: int,
    site_label: str,
    lon: str,
    lat: str,
    unit: str,
    age: np.ndarray,
    val: np.ndarray,
):
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    c1 = start_col
    c2 = start_col + 1

    ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
    ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)

    cell = ws.cell(row=1, column=c1, value=site_label)
    cell.font = bold
    cell.alignment = center

    coord = ""
    if lon or lat:
        coord = f"Lon={lon} , Lat={lat}"
    cell = ws.cell(row=2, column=c1, value=coord)
    cell.alignment = center

    cell = ws.cell(row=3, column=c1, value=unit)
    cell.alignment = center

    ws.cell(row=4, column=c1, value="Age").font = bold
    ws.cell(row=4, column=c2, value="Value").font = bold

    max_n = EXCEL_MAX_ROWS - 4
    n = min(len(age), len(val), max_n)
    for i in range(n):
        ws.cell(row=5 + i, column=c1, value=float(age[i]) if np.isfinite(age[i]) else "")
        ws.cell(row=5 + i, column=c2, value=float(val[i]) if np.isfinite(val[i]) else "")


# -------------------- Composite priority: filter by the same study site --------------------
def site_key(base_site: str, lon: str, lat: str) -> Tuple[str, str, str]:
    return (base_site or "", lon or "", lat or "")


def choose_series_for_same_site(series_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    If temperatureComposite exists, keep only one record (prioritize the one with more valid points, then column order).
    Otherwise, keep all records.
    """
    if not series_list:
        return []
    if not PREFER_TEMPERATURE_COMPOSITE:
        return series_list

    comps = [s for s in series_list if s.get("is_composite", False)]
    if comps:
        comps.sort(key=lambda x: (x.get("n_valid", 0), -x.get("col_idx0", 0)), reverse=True)
        return [comps[0]]

    series_list.sort(key=lambda x: (str(x.get("table_file", "")), int(x.get("col_idx0", 0))))
    return series_list


def main():
    meta = load_essential_metadata(METADATA_XLSX)

    if FILTER_ANNUAL_ONLY and meta["Season General"].astype(str).str.strip().eq("").all():
        raise ValueError("FILTER_ANNUAL_ONLY=True, but no usable 'Season General' column/value was found in the metadata. Turn off this switch or use the correct metadata file.")

    proxy_categories, by_dataset, by_site = build_proxy_general_lookup(meta)
    annual_by_dataset, annual_by_site, annual_by_dataset_pg, annual_by_site_pg = build_annual_lookup(meta)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    sheets: Dict[str, Any] = {}
    next_col: Dict[str, int] = {}
    for pg in proxy_categories:
        name = _safe_sheet_name(pg)
        ws = wb.create_sheet(name)
        sheets[pg] = ws
        next_col[pg] = 1

    site_name_global_counter: Dict[str, Dict[str, int]] = {pg: {} for pg in proxy_categories}

    lipd_files: List[str] = []
    for root, _, files in os.walk(LIPD_DIR):
        for fn in files:
            if not fn.lower().endswith(".lpd"):
                continue
            if fn.startswith("._"):
                continue
            fp = os.path.join(root, fn)
            if not zipfile.is_zipfile(fp):
                continue
            lipd_files.append(fp)
    lipd_files.sort()

    total = len(lipd_files)
    if total == 0:
        raise RuntimeError("No usable .lpd (zip) files were found. Please check the directory.")

    series_written = 0
    failed = 0
    no_temp = 0
    skipped_non_annual_files = 0
    skipped_non_annual_series = 0
    skipped_by_0_12k_coverage = 0

    unit_logs: List[Dict[str, Any]] = []

    for i, fp in enumerate(lipd_files, start=1):
        try:
            with zipfile.ZipFile(fp, "r") as zf:
                meta_json = _load_metadata_jsonld(zf)
                graph_idx = _build_graph_index(meta_json)
                dataset = _find_dataset_node(meta_json)
                dataset = _resolve_ref(dataset, graph_idx) if isinstance(dataset, dict) else meta_json

                dataset_name, site_name_lpd = _extract_dataset_and_site_name(dataset, graph_idx)
                lon, lat = _extract_lon_lat(dataset, graph_idx)

                # ===== Annual filtering (file level) =====
                if FILTER_ANNUAL_ONLY and (not is_annual_any(dataset_name, site_name_lpd, annual_by_dataset, annual_by_site)):
                    skipped_non_annual_files += 1
                    continue

                tables = _iter_measurement_tables(dataset, graph_idx)
                if not tables:
                    no_temp += 1
                    continue

                # Collect first: group by (pg, site_key) to support composite-priority selection
                collected: Dict[str, Dict[Tuple[str, str, str], List[Dict[str, Any]]]] = {}

                found_any_series_in_file = False

                for t in tables:
                    fname = _table_filename(t)
                    if not fname:
                        continue
                    cols_meta = _table_columns(t)
                    if not cols_meta:
                        continue

                    time_idx0 = pick_time_col(cols_meta)
                    if time_idx0 is None:
                        continue

                    temp_cols = find_temperature_columns(cols_meta)
                    if not temp_cols:
                        continue

                    df = _read_table_from_zip(zf, fname, _table_missing_value(t))
                    ncol = df.shape[1]
                    if not (0 <= time_idx0 < ncol):
                        continue

                    age_ser = pd.to_numeric(df.iloc[:, time_idx0], errors="coerce")

                    for idx0, units, blob, col_meta in temp_cols:
                        if not (0 <= idx0 < ncol):
                            continue

                        v_ser = pd.to_numeric(df.iloc[:, idx0], errors="coerce")
                        mask = age_ser.notna() & v_ser.notna()
                        n_valid = int(mask.sum())
                        if n_valid < 3:
                            continue

                        age = age_ser[mask].to_numpy(dtype=float)
                        val_raw = v_ser[mask].to_numpy(dtype=float)

                        # Proxy General
                        pg = choose_proxy_general(
                            dataset_name=dataset_name,
                            site_name=site_name_lpd,
                            blob=blob,
                            by_dataset=by_dataset,
                            by_site=by_site,
                            available=proxy_categories,
                        )

                        # ===== Annual filtering (series level) =====
                        if FILTER_ANNUAL_ONLY and (not is_annual_for_pg(dataset_name, site_name_lpd, pg, annual_by_dataset_pg, annual_by_site_pg)):
                            skipped_non_annual_series += 1
                            continue

                        # ===== Added: 0–12k 1-kyr coverage filter (series level) =====
                        if FILTER_0_12K_KYR_COVERAGE:
                            ok_cov, cov_info = check_kyr_coverage_0_12k(age)
                            if not ok_cov:
                                skipped_by_0_12k_coverage += 1
                                if WRITE_UNIT_LOG:
                                    unit_logs.append({
                                        "lpd_file": os.path.basename(fp),
                                        "dataset": dataset_name,
                                        "site": site_name_lpd,
                                        "lon": lon,
                                        "lat": lat,
                                        "proxy_general_assigned": pg,
                                        "table_file": fname,
                                        "col_idx0": idx0,
                                        "variableName": _col_name(col_meta),
                                        "units_raw": units,
                                        "unit_out": "",
                                        "is_temperatureComposite": "",
                                        "site_label_written": "",
                                        "note": (
                                            f"[SKIP_BY_0_12K_COVERAGE] reason={cov_info.get('reason')}; "
                                            f"empty_bins={cov_info.get('empty_bins')}; "
                                            f"max_consec_empty={cov_info.get('max_consecutive_empty')}; "
                                            f"counts={cov_info.get('counts')}"
                                        ),
                                        "blob": (blob or "")[:220],
                                        "n_valid": n_valid,
                                    })
                                continue

                        # Align units to degC / degC_anom
                        val, unit_show, note = normalize_temperature_to_degC(val_raw, units, blob)

                        base_site = site_name_lpd or dataset_name or os.path.splitext(os.path.basename(fp))[0]
                        sk = site_key(base_site, lon, lat)

                        varname = _col_name(col_meta)
                        is_comp = is_temperature_composite(varname, blob)

                        collected.setdefault(pg, {}).setdefault(sk, []).append({
                            "base_site": base_site,
                            "lon": lon,
                            "lat": lat,
                            "unit_show": unit_show,
                            "age": age,
                            "val": val,
                            "table_file": fname,
                            "col_idx0": idx0,
                            "variableName": varname,
                            "units_raw": units,
                            "blob": blob,
                            "note": note,
                            "n_valid": n_valid,
                            "is_composite": is_comp,
                            "lpd_file": os.path.basename(fp),
                            "dataset": dataset_name,
                            "site": site_name_lpd,
                            "pg": pg,
                        })

                        found_any_series_in_file = True

                if not found_any_series_in_file:
                    no_temp += 1
                    continue

                # ===== Apply composite-priority selection within the same study site and write to workbook =====
                for pg, site_dict in collected.items():
                    ws = sheets.get(pg)
                    if ws is None:
                        continue

                    for _, series_list in site_dict.items():
                        chosen = choose_series_for_same_site(series_list)
                        if not chosen:
                            continue

                        # Handle global uniqueness of base_site (same name but different coordinates within the same sheet)
                        base_site = chosen[0]["base_site"]
                        cnt_site = site_name_global_counter[pg].get(base_site, 0) + 1
                        site_name_global_counter[pg][base_site] = cnt_site
                        base_site_unique = base_site if cnt_site == 1 else f"{base_site}_S{cnt_site}"

                        multi = len(chosen) > 1

                        for j, s in enumerate(chosen, start=1):
                            site_label = base_site_unique if (not multi) else f"{base_site_unique}_{j}"

                            col = next_col[pg]
                            if col + 1 > EXCEL_MAX_COLS:
                                raise RuntimeError(f"Sheet '{pg}' columns exceeded Excel limit.")

                            write_series_block(
                                ws=ws,
                                start_col=col,
                                site_label=site_label,
                                lon=s["lon"],
                                lat=s["lat"],
                                unit=s["unit_show"],
                                age=s["age"],
                                val=s["val"],
                            )
                            next_col[pg] += 2
                            series_written += 1

                            if WRITE_UNIT_LOG:
                                unit_logs.append({
                                    "lpd_file": s["lpd_file"],
                                    "dataset": s["dataset"],
                                    "site": s["site"],
                                    "lon": s["lon"],
                                    "lat": s["lat"],
                                    "proxy_general_assigned": pg,
                                    "table_file": s["table_file"],
                                    "col_idx0": s["col_idx0"],
                                    "variableName": s["variableName"],
                                    "units_raw": s["units_raw"],
                                    "unit_out": s["unit_show"],
                                    "is_temperatureComposite": bool(s["is_composite"]),
                                    "site_label_written": site_label,
                                    "note": s["note"],
                                    "blob": (s["blob"] or "")[:220],
                                    "n_valid": s["n_valid"],
                                })

        except Exception:
            failed += 1

        if i % 100 == 0 or i == total:
            print(
                f"[PROGRESS] {i}/{total} | series_written={series_written} | failed={failed} | "
                f"no_temp={no_temp} | skipped_non_annual_files={skipped_non_annual_files} | "
                f"skipped_non_annual_series={skipped_non_annual_series} | "
                f"skipped_by_0_12k_coverage={skipped_by_0_12k_coverage}"
            )

    out_name = datetime.now().strftime("%Y-%m-%d_%H%M") + "_Temp12k_Temperature_TRUE_withCompositePriority.xlsx"
    out_path = os.path.join(SCRIPT_DIR, out_name)
    wb.save(out_path)

    print(f"[DONE] Saved: {out_path}")
    print(
        f"[SUMMARY] valid_lpd={total}, series_written={series_written}, failed_files={failed}, no_temp_files={no_temp}, "
        f"skipped_non_annual_files={skipped_non_annual_files}, skipped_non_annual_series={skipped_non_annual_series}, "
        f"skipped_by_0_12k_coverage={skipped_by_0_12k_coverage}, "
        f"FILTER_ANNUAL_ONLY={FILTER_ANNUAL_ONLY}, PREFER_TEMPERATURE_COMPOSITE={PREFER_TEMPERATURE_COMPOSITE}, "
        f"FILTER_0_12K_KYR_COVERAGE={FILTER_0_12K_KYR_COVERAGE}"
    )

    if WRITE_UNIT_LOG and unit_logs:
        pd.DataFrame(unit_logs).to_csv(SCRIPT_DIR / UNIT_LOG_CSV, index=False, encoding="utf-8-sig")
        print(f"[LOG] Unit log saved: {SCRIPT_DIR / UNIT_LOG_CSV}")


if __name__ == "__main__":
    main()
