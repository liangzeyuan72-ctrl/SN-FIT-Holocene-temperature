import pandas as pd
import numpy as np
import xarray as xr
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from pathlib import Path
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Alignment

# === Added: read the shapefile and build a land/ocean mask ===
from cartopy.io.shapereader import Reader
from shapely.ops import unary_union
from shapely.prepared import prep
from shapely.geometry import Point


# Configuration
CFG = {
    "DATA_DIR": Path(__file__).resolve().parent,
    "OUT_DIR": Path(__file__).resolve().parent / "correlation_trace_sn_vs_global_no_p",
    "VAR": "TS",
    "AGE_RANGE_KA": (0.0, 12.0),
    "SCALE_YEARS": 1000,  # Millennial resolution
    "FORCINGS": ["ghg-only", "orb-only", "ice-only", "fwf-only"],
    "FORCING_NAMES": ["GHG", "ORB", "ICE", "FWF"],
    "SN_LAT": 28.7,
    "SN_LON": 117.25,
}


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)
    return p


def get_script_dir() -> Path:
    # Use the script directory as the default location
    if "__file__" in globals():
        return Path(__file__).resolve().parent
    return Path.cwd()


def time_to_age_kaBP(time_kaBP: xr.DataArray):
    return -time_kaBP


def select_time_window(ds: xr.Dataset, age_min: float, age_max: float):
    t = ds["time"]
    tmin = -age_max
    tmax = -age_min
    m = (t >= tmin) & (t <= tmax) & (t <= 0.0)
    return ds.sel(time=m)


def add_year_coord(ds: xr.Dataset):
    age_ka = time_to_age_kaBP(ds["time"])
    age_year = age_ka * 1000.0
    year_bp = np.floor(age_year.values + 1e-9).astype(np.int32)
    return ds.assign_coords(age_kaBP=("time", age_ka.values), year_bp=("time", year_bp))


def monthly_to_annual_mean(da: xr.DataArray):
    annual = da.groupby("year_bp").mean("time")
    annual = annual.assign_coords(age_kaBP=("year_bp", annual["year_bp"].values.astype(np.float64) / 1000.0))
    annual = annual.swap_dims({"year_bp": "age_kaBP"})
    annual = annual.sortby("age_kaBP")
    return annual


def average_to_millennial(annual_da: xr.DataArray):
    scale_years = CFG["SCALE_YEARS"]
    age = annual_da["age_kaBP"].values
    bin_edges = np.arange(0, 12 + scale_years / 1000.0, scale_years / 1000.0)
    bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2

    binned = []
    for i in range(len(bin_centers)):
        mask = (age >= bin_edges[i]) & (age < bin_edges[i + 1])
        if mask.sum() > 0:
            mean_slice = annual_da.isel(age_kaBP=mask).mean(dim="age_kaBP")
            binned.append(mean_slice)
        else:
            dummy = annual_da.isel(age_kaBP=0).copy()
            dummy[:] = np.nan
            binned.append(dummy)

    da_binned = xr.concat(binned, dim="age_kaBP")
    da_binned = da_binned.assign_coords(age_kaBP=("age_kaBP", bin_centers))
    return da_binned


def calculate_correlation_areas(corr_map, lat):
    """
    Your original global (land + ocean) area-fraction statistics are kept unchanged
    """
    if np.all(np.isnan(corr_map)):
        return {
            "Global_Positive": np.nan,
            "Global_Negative": np.nan,
            "Lat30_Positive": np.nan,
            "Lat30_Negative": np.nan,
            "Lat40_Positive": np.nan,
            "Lat40_Negative": np.nan,
            "NH_Positive": np.nan,
            "NH_Negative": np.nan,
            "SH_Positive": np.nan,
            "SH_Negative": np.nan,
        }

    weights = np.cos(np.deg2rad(lat))
    weights = np.repeat(weights[:, np.newaxis], corr_map.shape[1], axis=1)

    total_weight_global = np.nansum(weights[~np.isnan(corr_map)])
    global_positive = np.nansum(weights[corr_map > 0]) / total_weight_global * 100 if total_weight_global > 0 else np.nan
    global_negative = np.nansum(weights[corr_map < 0]) / total_weight_global * 100 if total_weight_global > 0 else np.nan

    lat_mask_30 = (lat >= -30) & (lat <= 30)
    corr_map_30 = corr_map[lat_mask_30, :]
    weights_30 = weights[lat_mask_30, :]
    total_weight_30 = np.nansum(weights_30[~np.isnan(corr_map_30)])
    lat30_positive = np.nansum(weights_30[corr_map_30 > 0]) / total_weight_30 * 100 if total_weight_30 > 0 else np.nan
    lat30_negative = np.nansum(weights_30[corr_map_30 < 0]) / total_weight_30 * 100 if total_weight_30 > 0 else np.nan

    lat_mask_40 = (lat >= -40) & (lat <= 40)
    corr_map_40 = corr_map[lat_mask_40, :]
    weights_40 = weights[lat_mask_40, :]
    total_weight_40 = np.nansum(weights_40[~np.isnan(corr_map_40)])
    lat40_positive = np.nansum(weights_40[corr_map_40 > 0]) / total_weight_40 * 100 if total_weight_40 > 0 else np.nan
    lat40_negative = np.nansum(weights_40[corr_map_40 < 0]) / total_weight_40 * 100 if total_weight_40 > 0 else np.nan

    lat_mask_nh = lat > 0
    corr_map_nh = corr_map[lat_mask_nh, :]
    weights_nh = weights[lat_mask_nh, :]
    total_weight_nh = np.nansum(weights_nh[~np.isnan(corr_map_nh)])
    nh_positive = np.nansum(weights_nh[corr_map_nh > 0]) / total_weight_nh * 100 if total_weight_nh > 0 else np.nan
    nh_negative = np.nansum(weights_nh[corr_map_nh < 0]) / total_weight_nh * 100 if total_weight_nh > 0 else np.nan

    lat_mask_sh = lat < 0
    corr_map_sh = corr_map[lat_mask_sh, :]
    weights_sh = weights[lat_mask_sh, :]
    total_weight_sh = np.nansum(weights_sh[~np.isnan(corr_map_sh)])
    sh_positive = np.nansum(weights_sh[corr_map_sh > 0]) / total_weight_sh * 100 if total_weight_sh > 0 else np.nan
    sh_negative = np.nansum(weights_sh[corr_map_sh < 0]) / total_weight_sh * 100 if total_weight_sh > 0 else np.nan

    return {
        "Global_Positive": round(global_positive, 1),
        "Global_Negative": round(global_negative, 1),
        "Lat30_Positive": round(lat30_positive, 1),
        "Lat30_Negative": round(lat30_negative, 1),
        "Lat40_Positive": round(lat40_positive, 1),
        "Lat40_Negative": round(lat40_negative, 1),
        "NH_Positive": round(nh_positive, 1),
        "NH_Negative": round(nh_negative, 1),
        "SH_Positive": round(sh_positive, 1),
        "SH_Negative": round(sh_negative, 1),
    }


# === Added: read shp / build land mask / compute area fractions within a mask ===
def load_land_geometry(shp_path: Path):
    reader = Reader(str(shp_path))
    geoms = list(reader.geometries())
    if len(geoms) == 0:
        raise RuntimeError(f"No geometries were found in the shapefile: {shp_path}")
    land_geom = unary_union(geoms)
    return land_geom, prep(land_geom)


def build_land_mask(lat_1d: np.ndarray, lon_1d: np.ndarray, land_geom, land_prepared):
    """
    True = land, False = ocean
    Convert 0–360 longitudes to -180–180 (to match common shapefile coordinates)
    """
    lon2d, lat2d = np.meshgrid(lon_1d, lat_1d)
    lon2d_180 = ((lon2d + 180.0) % 360.0) - 180.0

    # Prefer shapely 2.x contains_xy (avoids deprecation warnings)
    try:
        from shapely import contains_xy  # shapely>=2.0
        return contains_xy(land_geom, lon2d_180, lat2d)
    except Exception:
        # shapely 1.x vectorized.contains
        try:
            from shapely import vectorized
            return vectorized.contains(land_geom, lon2d_180, lat2d)
        except Exception:
            # Fallback: point-by-point evaluation (slow but robust)
            nlat, nlon = lon2d_180.shape
            out = np.zeros((nlat, nlon), dtype=bool)
            for j in range(nlat):
                for k in range(nlon):
                    out[j, k] = land_prepared.contains(Point(float(lon2d_180[j, k]), float(lat2d[j, k])))
            return out


def calculate_correlation_areas_masked(corr_map: np.ndarray, lat: np.ndarray, mask2d: np.ndarray):
    """
    Revised version: all region_valid arrays keep shape (nlat, nlon) without slicing, avoiding boolean-index dimension mismatches.
    Within the grid cells where mask2d=True, compute positive/negative correlation area fractions for each region (weighted by cos(lat)).
    """
    if np.all(np.isnan(corr_map)):
        return {
            "Global_Positive": np.nan,
            "Global_Negative": np.nan,
            "Lat30_Positive": np.nan,
            "Lat30_Negative": np.nan,
            "Lat40_Positive": np.nan,
            "Lat40_Negative": np.nan,
            "NH_Positive": np.nan,
            "NH_Negative": np.nan,
            "SH_Positive": np.nan,
            "SH_Negative": np.nan,
        }

    lat = np.asarray(lat)
    nlat, nlon = corr_map.shape

    # weights: (nlat,nlon)
    w_lat = np.cos(np.deg2rad(lat))[:, None]
    weights = np.repeat(w_lat, nlon, axis=1)

    # valid: (nlat,nlon)
    valid = (mask2d.astype(bool)) & (~np.isnan(corr_map))

    pos = corr_map > 0
    neg = corr_map < 0

    def pct(sign_mask_2d, region_valid_2d):
        tw = np.nansum(weights[region_valid_2d])
        if tw <= 0:
            return np.nan
        return np.nansum(weights[sign_mask_2d & region_valid_2d]) / tw * 100.0

    def region_from_latmask(latmask_1d: np.ndarray):
        return valid & (latmask_1d[:, None])

    # Global (within mask)
    global_pos = pct(pos, valid)
    global_neg = pct(neg, valid)

    # -30~30（within mask）
    latmask_30 = (lat >= -30) & (lat <= 30)
    reg30 = region_from_latmask(latmask_30)
    lat30_pos = pct(pos, reg30)
    lat30_neg = pct(neg, reg30)

    # -40~40（within mask）
    latmask_40 = (lat >= -40) & (lat <= 40)
    reg40 = region_from_latmask(latmask_40)
    lat40_pos = pct(pos, reg40)
    lat40_neg = pct(neg, reg40)

    # NH（within mask）
    latmask_nh = lat > 0
    regnh = region_from_latmask(latmask_nh)
    nh_pos = pct(pos, regnh)
    nh_neg = pct(neg, regnh)

    # SH（within mask）
    latmask_sh = lat < 0
    regsh = region_from_latmask(latmask_sh)
    sh_pos = pct(pos, regsh)
    sh_neg = pct(neg, regsh)

    def r(x):
        return np.nan if np.isnan(x) else round(x, 1)

    return {
        "Global_Positive": r(global_pos),
        "Global_Negative": r(global_neg),
        "Lat30_Positive": r(lat30_pos),
        "Lat30_Negative": r(lat30_neg),
        "Lat40_Positive": r(lat40_pos),
        "Lat40_Negative": r(lat40_neg),
        "NH_Positive": r(nh_pos),
        "NH_Negative": r(nh_neg),
        "SH_Positive": r(sh_pos),
        "SH_Negative": r(sh_neg),
    }


def write_summary_sheet(ws, df_areas: pd.DataFrame):
    regions = ["Global", "Lat -30~30", "Lat -40~40", "Northern Hemisphere", "Southern Hemisphere"]
    col = 2
    ws.cell(row=1, column=1, value="Forcing")
    for region in regions:
        ws.cell(row=1, column=col, value=region)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col, value="Positive")
        ws.cell(row=2, column=col + 1, value="Negative")
        col += 2

    for r_i, (forcing, row_data) in enumerate(df_areas.iterrows(), start=3):
        ws.cell(row=r_i, column=1, value=forcing)
        ws.cell(row=r_i, column=2, value=row_data["Global_Positive"])
        ws.cell(row=r_i, column=3, value=row_data["Global_Negative"])
        ws.cell(row=r_i, column=4, value=row_data["Lat30_Positive"])
        ws.cell(row=r_i, column=5, value=row_data["Lat30_Negative"])
        ws.cell(row=r_i, column=6, value=row_data["Lat40_Positive"])
        ws.cell(row=r_i, column=7, value=row_data["Lat40_Negative"])
        ws.cell(row=r_i, column=8, value=row_data["NH_Positive"])
        ws.cell(row=r_i, column=9, value=row_data["NH_Negative"])
        ws.cell(row=r_i, column=10, value=row_data["SH_Positive"])
        ws.cell(row=r_i, column=11, value=row_data["SH_Negative"])


def main():
    out_dir = ensure_dir(CFG["OUT_DIR"])
    print(f"Output folder: {out_dir}")

    # Look for World_countries.shp in the script directory
    script_dir = get_script_dir()
    shp_path = script_dir / "World_countries.shp"
    if not shp_path.exists():
        raise FileNotFoundError(
            f"Could not find  {shp_path}\n"
            f"Place World_countries.shp together with its .shx/.dbf/.prj companion files in the script directory."
        )
    land_geom, land_prepared = load_land_geometry(shp_path)

    results_list = []        # All (land + ocean, original logic)
    results_land_list = []   # Land
    results_ocean_list = []  # Ocean

    land_mask_cache = None   # (nlat, nlon), built only once

    for forcing_file, forcing_name in zip(CFG["FORCINGS"], CFG["FORCING_NAMES"]):
        fpath = CFG["DATA_DIR"] / f"TraCE-21K-{forcing_file}.monthly.TS.nc"
        if not fpath.exists():
            print(f"File not found: {fpath}")
            continue

        print(f"\n=== Processing {forcing_name} ===")
        ds = xr.open_dataset(fpath, decode_times=False)

        ds2 = select_time_window(ds, *CFG["AGE_RANGE_KA"])
        ds3 = add_year_coord(ds2)
        annual = monthly_to_annual_mean(ds3[CFG["VAR"]])

        # Millennial averaging
        binned_da = average_to_millennial(annual)

        # Build the land mask (cached and reused)
        if land_mask_cache is None:
            land_mask_cache = build_land_mask(
                binned_da["lat"].values,
                binned_da["lon"].values,
                land_geom,
                land_prepared
            )
            print(f"  Land mask built: shape={land_mask_cache.shape} (True=Land, False=Ocean)")

        # Extract the SN point series (nearest grid cell)
        sn_point_seq = binned_da.sel(lat=CFG["SN_LAT"], lon=CFG["SN_LON"], method="nearest").values

        # Compute the correlation map (without p values)
        nlat = binned_da.sizes["lat"]
        nlon = binned_da.sizes["lon"]
        corr_map = np.full((nlat, nlon), np.nan)

        for j in range(nlat):
            for k in range(nlon):
                grid_seq = binned_da.isel(lat=j, lon=k).values
                m = ~np.isnan(grid_seq) & ~np.isnan(sn_point_seq)
                if m.sum() < 5:
                    continue
                corr, _ = pearsonr(grid_seq[m], sn_point_seq[m])
                corr_map[j, k] = corr

        # Saved NetCDF
        da_corr = xr.DataArray(
            corr_map,
            dims=("lat", "lon"),
            coords={"lat": binned_da["lat"], "lon": binned_da["lon"]},
            name="correlation",
        )
        out_nc = out_dir / f"{forcing_name}_millennial.nc"
        da_corr.to_dataset().to_netcdf(out_nc)
        print(f"  Saved NetCDF: {out_nc}")

        # Save the CSV data table
        lat_flat = np.repeat(binned_da["lat"].values, nlon)
        lon_flat = np.tile(binned_da["lon"].values, nlat)
        corr_flat = corr_map.flatten()
        df_table = pd.DataFrame({"lat": lat_flat, "lon": lon_flat, "correlation": corr_flat})
        out_csv = out_dir / f"{forcing_name}_millennial_data.csv"
        df_table.to_csv(out_csv, index=False)
        print(f"  Saved data table: {out_csv}")

        # Area fractions: All (original) + Land/Ocean (added)
        areas_all = calculate_correlation_areas(corr_map, binned_da["lat"].values)
        areas_all["Forcing"] = forcing_name
        results_list.append(areas_all)

        areas_land = calculate_correlation_areas_masked(corr_map, binned_da["lat"].values, land_mask_cache)
        areas_land["Forcing"] = forcing_name
        results_land_list.append(areas_land)

        areas_ocean = calculate_correlation_areas_masked(corr_map, binned_da["lat"].values, ~land_mask_cache)
        areas_ocean["Forcing"] = forcing_name
        results_ocean_list.append(areas_ocean)

        # Plot (unchanged)
        fig = plt.figure(figsize=(10, 5))
        ax = fig.add_subplot(1, 1, 1, projection=ccrs.PlateCarree())
        im = ax.pcolormesh(
            binned_da["lon"],
            binned_da["lat"],
            corr_map,
            cmap="RdBu_r",
            vmin=-1,
            vmax=1,
            transform=ccrs.PlateCarree(),
        )
        ax.add_feature(cfeature.COASTLINE)
        ax.set_title(f"{forcing_name} - Trace SN Point vs Global Correlation (Millennial)", fontsize=12)

        fig.colorbar(
            im,
            ax=ax,
            orientation="vertical",
            label="Pearson r",
            shrink=0.45,
            pad=0.02,
            aspect=30,
            fraction=0.046,
        )

        fig.subplots_adjust(top=0.92, bottom=0.10, left=0.06, right=0.88, wspace=0.02, hspace=0.02)

        out_png = out_dir / f"{forcing_name}_millennial_sn_corr.png"
        plt.savefig(out_png, dpi=300, bbox_inches="tight")
        plt.close()
        print(f"  Saved figure: {out_png}")

    # Excel workbook with three sheets: All / Land / Ocean
    df_all = pd.DataFrame(results_list).set_index("Forcing")
    df_land = pd.DataFrame(results_land_list).set_index("Forcing")
    df_ocean = pd.DataFrame(results_ocean_list).set_index("Forcing")

    out_table = out_dir / "areas_summary.xlsx"
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "All"
    write_summary_sheet(ws_all, df_all)

    ws_land = wb.create_sheet("Land")
    write_summary_sheet(ws_land, df_land)

    ws_ocean = wb.create_sheet("Ocean")
    write_summary_sheet(ws_ocean, df_ocean)

    wb.save(out_table)

    print(f"\nArea-fraction summary table: {out_table}")
    print("\n[All]")
    print(df_all.to_string())
    print("\n[Land]")
    print(df_land.to_string())
    print("\n[Ocean]")
    print(df_ocean.to_string())

    print("\n[Done]")
    print("Outputs saved in:", out_dir)


if __name__ == "__main__":
    main()