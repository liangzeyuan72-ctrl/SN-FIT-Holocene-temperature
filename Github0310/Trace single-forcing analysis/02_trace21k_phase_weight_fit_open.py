# -*- coding: utf-8 -*-
import os
import re
import time
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from scipy.optimize import minimize
from statsmodels.nonparametric.smoothers_lowess import lowess


# ============================================================
# 0) Edit here only if needed: default directory
# ============================================================
BASE_DIR = Path(__file__).resolve().parent

# Raw TraCE point file — used to build loess0.2 if the compact loess file is absent
TRACE_RAW_FILE = BASE_DIR / "TraCE21K_TS_point_28.7N_117.2E_annual_merged.xlsx"

# Compact loess0.2 file (used first if present; fastest option)
TRACE_LOESS_COMPACT = BASE_DIR / "TraCE21K_TS_point_28.7N_117.2E_annual_merged_loess0.20_compact.xlsx"

# SN file
SN_FILE = BASE_DIR / "SN-FIT BG.xlsx"

# Output: SN_MIX4_TRACE_YYYYMMDD.xlsx
OUT_XLSX = BASE_DIR / f"0119 ORB SN_MIX4_TRACE_{datetime.now().strftime('%Y%m%d')}.xlsx"


# ============================================================
# 1) Core settings
# ============================================================
AGE_MIN, AGE_MAX = 0.0, 11.2

# Shared lag search
LAG_RANGE = (0.0, 0.5)
LAG_STEP  = 0.01   # requested 0.01

# ORB weight scan (method 1)
ORB_RANGE = (0.01, 0.95)
ORB_STEP  = 0.01   # requested 0.01

# Lower bound for the remaining weights (GHG/ICE/FWF in each stage >= MIN_W)
MIN_W = 0.01

# Turning-point windows (automatically identified from SN)
WIN_LATE_MAX = (1.0, 1.5)     # Late-Holocene maximum (used as a constraint, not as a stage boundary)
WIN_LATE_MIN = (3.5, 4.5)     # Low around 4 ka (boundary 1)
WIN_MID_MAX  = (6.0, 7.0)     # Peak at 6–7 ka (boundary 2)
WIN_EARLY_MIN = (9.5, 11.2)   # Early-Holocene minimum (constraint)

# Loss-function weights (tunable; this set usually keeps the mid-Holocene peak within 6–7 ka)
ALPHA_ANCHOR = 6.0     # Anchor-matching weight (fit peak/trough positions and relative amplitude)
BETA_PEAK    = 8.0     # Penalty for an overly high late peak relative to the mid peak
LAMBDA_CONT  = 12.0    # Continuity penalty (avoid jumps at stage boundaries)
GAMMA_SMOOTH = 0.6     # Stage-weight smoothness penalty (avoid abrupt coefficient changes)

# "1% rule": within global_best_loss * (1+EPS), choose the largest w_ORB
EPS_1PCT = 0.01


# ============================================================
# 2) General utility functions
# ============================================================
def _as_numeric(a):
    return pd.to_numeric(a, errors="coerce").to_numpy(dtype=float)

def zscore_nan(x):
    x = np.asarray(x, float)
    m = np.isfinite(x)
    if m.sum() < 10:
        return np.full_like(x, np.nan, dtype=float)
    mu = float(np.nanmean(x[m]))
    sd = float(np.nanstd(x[m]))
    if (not np.isfinite(sd)) or sd == 0:
        sd = 1.0
    return (x - mu) / sd

def interp_nan(x_src, y_src, x_tgt):
    """Interpolate only within the source range; assign NaN outside to avoid endpoint extrapolation."""
    x_src = np.asarray(x_src, float)
    y_src = np.asarray(y_src, float)
    x_tgt = np.asarray(x_tgt, float)

    m = np.isfinite(x_src) & np.isfinite(y_src)
    if m.sum() < 5:
        return np.full_like(x_tgt, np.nan, dtype=float)

    xs = x_src[m]
    ys = y_src[m]
    order = np.argsort(xs)
    xs = xs[order]
    ys = ys[order]

    out = np.interp(x_tgt, xs, ys)
    out[(x_tgt < xs.min()) | (x_tgt > xs.max())] = np.nan
    return out

def interp_shift(age_src, val_src, age_tgt, lag_ka):
    """
    lag > 0 shifts the series toward older ages (response lag), then interpolates onto age_tgt.
    """
    return interp_nan(np.asarray(age_src) + float(lag_ka), val_src, age_tgt)

def window_argextreme(age, series, win, mode="max"):
    a, b = win
    age = np.asarray(age, float)
    series = np.asarray(series, float)
    idx = np.where((age >= a) & (age <= b) & np.isfinite(series))[0]
    if idx.size == 0:
        return None
    s = series[idx]
    if mode == "max":
        return float(age[idx[int(np.nanargmax(s))]])
    else:
        return float(age[idx[int(np.nanargmin(s))]])

def value_at(age, series, target_age):
    age = np.asarray(age, float)
    series = np.asarray(series, float)
    m = np.isfinite(age) & np.isfinite(series)
    if m.sum() < 10:
        return np.nan
    a = age[m]
    s = series[m]
    if target_age < a.min() or target_age > a.max():
        return np.nan
    order = np.argsort(a)
    a = a[order]
    s = s[order]
    return float(np.interp(target_age, a, s))

def safe_out_path(path: Path) -> Path:
    """Avoid PermissionError: if the file already exists and is in use, create a new filename automatically."""
    if not path.exists():
        return path
    stem = path.stem
    suf = path.suffix
    for k in range(1, 1000):
        p2 = path.with_name(f"{stem}_{k:03d}{suf}")
        if not p2.exists():
            return p2
    return path.with_name(f"{stem}_{int(time.time())}{suf}")


# ============================================================
# 3) Read SN
# ============================================================
def read_sn_series(path: Path):
    df = pd.read_excel(path)
    if df.shape[1] < 2:
        df = pd.read_excel(path, header=None)

    age_col = df.columns[0]
    val_col = df.columns[1]

    for c in df.columns:
        s = str(c).lower()
        if any(k in s for k in ["age", "bp", "yr", "year", "ka"]):
            age_col = c
            break

    for c in df.columns:
        s = str(c).lower()
        if any(k in s for k in ["sn", "fit", "bg", "temp", "ts"]):
            if c != age_col:
                val_col = c
                break

    age = _as_numeric(df[age_col])
    val = _as_numeric(df[val_col])
    m = np.isfinite(age) & np.isfinite(val)
    age = age[m]
    val = val[m]

    # SN is in a BP (years); convert to ka BP
    if np.nanmax(age) > 200:
        age_ka = age / 1000.0
        unit = "a BP -> ka BP (div1k)"
    else:
        age_ka = age
        unit = "ka BP (as is)"

    # Trim to 0–11.2 ka
    m2 = (age_ka >= AGE_MIN) & (age_ka <= AGE_MAX)
    age_ka = age_ka[m2]
    val = val[m2]

    order = np.argsort(age_ka)
    return age_ka[order], val[order], unit


# ============================================================
# 4) Read TraCE forcings (prefer compact loess; otherwise build loess from the raw file)
# ============================================================
def detect_time_col(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    # Priority
    priority = ["time_kyr", "time", "year", "yr", "age"]
    for p in priority:
        for c in cols:
            if p in str(c).lower():
                return c
    # Fallback: the first column that can be converted to numeric
    for c in cols:
        x = pd.to_numeric(df[c], errors="coerce")
        if np.isfinite(x).sum() > 50:
            return c
    return cols[0]

def time_to_age_ka(t):
    t = np.asarray(t, float)
    # Common case: TraCE time_kyr is negative and denotes the past; convert with age_kaBP = -time_kyr
    if np.nanmedian(t[np.isfinite(t)]) < 0:
        age = -t
        rule = "time_kyr (neg) -> age_kaBP = -time"
    else:
        age = t
        rule = "time already positive -> age_kaBP = time"
    # If the unit looks like years, divide by 1000 (normally not needed for TraCE)
    if np.nanmedian(np.abs(age[np.isfinite(age)])) > 200:
        age = age / 1000.0
        rule += " | unit=yr_like div1k"
    else:
        rule += " | unit=kyr_like"
    return age, rule

def detect_forcing_cols(df: pd.DataFrame, require_loess=True):
    """
    Identify the four columns: GHG/ORB/ICE/FWF
    Compact loess files usually contain loess02 in the column names
    """
    cols = list(df.columns)
    low = {c: str(c).lower() for c in cols}

    def pick(keywords):
        cand = []
        for c in cols:
            s = low[c]
            ok = all(k in s for k in keywords)
            if ok:
                if require_loess:
                    if ("loess" in s) or ("loess02" in s) or ("loess0.2" in s):
                        cand.append(c)
                else:
                    cand.append(c)
        if len(cand) == 0:
            return None
        # Prefer columns that look more like temperature series (contain ts or _c)
        cand2 = [c for c in cand if ("ts" in low[c] or "_c" in low[c])]
        return cand2[0] if len(cand2) else cand[0]

    col_GHG = pick(["ghg"])
    col_ORB = pick(["orb"])
    col_ICE = pick(["ice"])
    col_FWF = pick(["fwf"]) or pick(["fresh"]) or pick(["melt"])

    return {"GHG": col_GHG, "ORB": col_ORB, "ICE": col_ICE, "FWF": col_FWF}

def build_compact_loess_from_raw(raw_file: Path, out_compact: Path, frac=0.2):
    df = pd.read_excel(raw_file)

    time_col = detect_time_col(df)
    t = _as_numeric(df[time_col])
    age, rule = time_to_age_ka(t)

    # Forcing columns (the raw file may not contain loess columns)
    cols_map = detect_forcing_cols(df, require_loess=False)
    if any(cols_map[k] is None for k in ["GHG", "ORB", "ICE", "FWF"]):
        raise RuntimeError(f"[RAW] [RAW] Unable to identify the four forcing columns in the raw TraCE file: {cols_map}")

    # Trim to the target range (with small boundary padding to avoid poor loess edges)
    pad = max(abs(LAG_RANGE[0]), abs(LAG_RANGE[1])) + 0.2
    m = np.isfinite(age) & (age >= AGE_MIN - pad) & (age <= AGE_MAX + pad)
    age = age[m]

    out = pd.DataFrame({"time_col": t[m]})
    out["age_kaBP"] = age

    for key in ["GHG", "ORB", "ICE", "FWF"]:
        y = _as_numeric(df[cols_map[key]])[m]
        # lowess requires finite values
        mm = np.isfinite(age) & np.isfinite(y)
        xs = age[mm]
        ys = y[mm]
        order = np.argsort(xs)
        xs = xs[order]
        ys = ys[order]

        y_sm = lowess(ys, xs, frac=frac, it=1, return_sorted=False)
        out[f"{key}_loess02"] = interp_nan(xs, y_sm, age)  # back to the full age series

    out_compact = safe_out_path(out_compact)
    out.to_excel(out_compact, index=False)
    return out_compact, time_col, rule, cols_map

def read_trace_loess(compact_file: Path):
    df = pd.read_excel(compact_file)

    # Compatible with earlier compact files: the time column may be time_kyr or year_like
    time_col = detect_time_col(df)
    t = _as_numeric(df[time_col])
    age, rule = time_to_age_ka(t)

    cols_map = detect_forcing_cols(df, require_loess=True)

    # If compact column names are not in the standard form, allow a fallback (e.g., GHG_loess02)
    for k in cols_map:
        if cols_map[k] is None:
            # Try finding "{k}_loess02" directly
            alt = [c for c in df.columns if f"{k.lower()}_loess02" in str(c).lower()]
            cols_map[k] = alt[0] if len(alt) else None

    if any(cols_map[k] is None for k in ["GHG", "ORB", "ICE", "FWF"]):
        raise RuntimeError(f"[COMPACT] [COMPACT] Unable to identify the four loess0.2 columns: {cols_map}")

    ghg = _as_numeric(df[cols_map["GHG"]])
    orb = _as_numeric(df[cols_map["ORB"]])
    ice = _as_numeric(df[cols_map["ICE"]])
    fwf = _as_numeric(df[cols_map["FWF"]])

    # Trim to the target range (add padding to support lag interpolation)
    pad = max(abs(LAG_RANGE[0]), abs(LAG_RANGE[1])) + 0.2
    m = np.isfinite(age) & np.isfinite(ghg) & np.isfinite(orb) & np.isfinite(ice) & np.isfinite(fwf)
    age = age[m]; ghg = ghg[m]; orb = orb[m]; ice = ice[m]; fwf = fwf[m]
    m2 = (age >= AGE_MIN - pad) & (age <= AGE_MAX + pad)
    age = age[m2]; ghg = ghg[m2]; orb = orb[m2]; ice = ice[m2]; fwf = fwf[m2]

    order = np.argsort(age)
    return age[order], ghg[order], orb[order], ice[order], fwf[order], time_col, rule, cols_map


# ============================================================
# 5) Core fitting with three stages + ORB scan
# ============================================================
def piecewise_mix(age, Gz, Oz, Iz, Fz, worb, w_stage, a1, a2):
    """
    w_stage: (wG1,wI1,wF1,  wG2,wI2,wF2,  wG3,wI3,wF3)
    """
    wG1,wI1,wF1, wG2,wI2,wF2, wG3,wI3,wF3 = w_stage
    mix = np.full_like(age, np.nan, dtype=float)

    m1 = (age <= a1)
    m2 = (age > a1) & (age <= a2)
    m3 = (age > a2)

    mix[m1] = worb*Oz[m1] + wG1*Gz[m1] + wI1*Iz[m1] + wF1*Fz[m1]
    mix[m2] = worb*Oz[m2] + wG2*Gz[m2] + wI2*Iz[m2] + wF2*Fz[m2]
    mix[m3] = worb*Oz[m3] + wG3*Gz[m3] + wI3*Iz[m3] + wF3*Fz[m3]
    return mix

def make_loss_fn(age, SNz, Gz, Oz, Iz, Fz, worb, a1, a2,
                 a_late_max, a_late_min, a_mid_max, a_early_min,
                 peak_tol=0.15):
    """
    Return loss(x) for the optimizer
    x = 9 weights: (G1,I1,F1, G2,I2,F2, G3,I3,F3)
    """

    remain = 1.0 - worb

    def loss(x):
        x = np.asarray(x, float)
        w_stage = tuple(x.tolist())
        mix = piecewise_mix(age, Gz, Oz, Iz, Fz, worb, w_stage, a1, a2)

        m = np.isfinite(SNz) & np.isfinite(mix)
        if m.sum() < 30:
            return 1e9

        # 1) Full-interval SSE
        sse = float(np.nanmean((SNz[m] - mix[m])**2))

        # 2) Anchor penalty: four key windows (your peak-trough logic)
        anchors = [a_late_max, a_late_min, a_mid_max, a_early_min]
        pen_anchor = 0.0
        for aa in anchors:
            y_sn = value_at(age, SNz, aa)
            y_mx = value_at(age, mix, aa)
            if np.isfinite(y_sn) and np.isfinite(y_mx):
                pen_anchor += (y_sn - y_mx)**2

        # 3) Peak penalty: suppress a 1 ka peak that is much higher than the 6–7 ka peak
        y_late = value_at(age, mix, a_late_max)
        y_mid  = value_at(age, mix, a_mid_max)
        pen_peak = 0.0
        if np.isfinite(y_late) and np.isfinite(y_mid):
            # If the late peak is too much higher than the mid peak (exceeding peak_tol), apply a penalty
            exceed = (y_late - y_mid) - peak_tol
            if exceed > 0:
                pen_peak = exceed**2

        # 4) Continuity penalty: avoid jumps at stage boundaries
        # Evaluate the same boundary age with two adjacent stage weights and penalize the difference
        # Boundary 1: stage 1 vs stage 2
        mix_b1_1 = worb*value_at(age, Oz, a1) + x[0]*value_at(age, Gz, a1) + x[1]*value_at(age, Iz, a1) + x[2]*value_at(age, Fz, a1)
        mix_b1_2 = worb*value_at(age, Oz, a1) + x[3]*value_at(age, Gz, a1) + x[4]*value_at(age, Iz, a1) + x[5]*value_at(age, Fz, a1)

        # Boundary 2: stage 2 vs stage 3
        mix_b2_2 = worb*value_at(age, Oz, a2) + x[3]*value_at(age, Gz, a2) + x[4]*value_at(age, Iz, a2) + x[5]*value_at(age, Fz, a2)
        mix_b2_3 = worb*value_at(age, Oz, a2) + x[6]*value_at(age, Gz, a2) + x[7]*value_at(age, Iz, a2) + x[8]*value_at(age, Fz, a2)

        pen_cont = 0.0
        if np.isfinite(mix_b1_1) and np.isfinite(mix_b1_2):
            pen_cont += (mix_b1_1 - mix_b1_2)**2
        if np.isfinite(mix_b2_2) and np.isfinite(mix_b2_3):
            pen_cont += (mix_b2_2 - mix_b2_3)**2

        # 5) Stage smoothness penalty: keep stage weights from changing too abruptly
        w1 = x[0:3]; w2 = x[3:6]; w3 = x[6:9]
        pen_smooth = float(np.sum((w1-w2)**2) + np.sum((w2-w3)**2))

        # Total loss
        return sse + ALPHA_ANCHOR*pen_anchor + BETA_PEAK*pen_peak + LAMBDA_CONT*pen_cont + GAMMA_SMOOTH*pen_smooth

    return loss

def optimize_for_fixed_worb_lag(age, SNz, Gz, Oz, Iz, Fz, worb, a1, a2,
                                a_late_max, a_late_min, a_mid_max, a_early_min,
                                x0=None):
    remain = 1.0 - worb

    # x is 9-D: 3 stages * (G, I, F)
    if x0 is None:
        x0 = np.array([remain/3]*9, dtype=float)

    bounds = [(MIN_W, remain) for _ in range(9)]

    # For each stage, sum = remain
    cons = [
        {"type": "eq", "fun": lambda x: (x[0]+x[1]+x[2]) - remain},
        {"type": "eq", "fun": lambda x: (x[3]+x[4]+x[5]) - remain},
        {"type": "eq", "fun": lambda x: (x[6]+x[7]+x[8]) - remain},
    ]

    loss_fn = make_loss_fn(
        age, SNz, Gz, Oz, Iz, Fz, worb, a1, a2,
        a_late_max=a_late_max,
        a_late_min=a_late_min,
        a_mid_max=a_mid_max,
        a_early_min=a_early_min,
        peak_tol=0.15
    )

    res = minimize(
        loss_fn,
        x0=x0,
        method="SLSQP",
        bounds=bounds,
        constraints=cons,
        options={"maxiter": 250, "ftol": 1e-7, "disp": False},
    )

    if not res.success:
        return None

    x = res.x
    # Compute mix and correlation coefficient
    mix = piecewise_mix(age, Gz, Oz, Iz, Fz, worb, tuple(x.tolist()), a1, a2)
    m = np.isfinite(SNz) & np.isfinite(mix)
    corr = float(np.corrcoef(SNz[m], mix[m])[0, 1]) if m.sum() > 30 else np.nan
    return {
        "loss": float(res.fun),
        "worb": float(worb),
        "w_stage": tuple(x.tolist()),
        "corr": corr,
        "mix_z": mix
    }


# ============================================================
# 6) Main program
# ============================================================
def main():
    if not SN_FILE.exists():
        raise FileNotFoundError(f"SN file not found: {SN_FILE}")

    sn_age, sn_val, sn_unit = read_sn_series(SN_FILE)
    sn_mean = float(np.nanmean(sn_val[np.isfinite(sn_val)]))
    sn_std  = float(np.nanstd(sn_val[np.isfinite(sn_val)]))
    sn_z = zscore_nan(sn_val)

    print(f"[INFO] SN age unit: {sn_unit}")

    # ---- TraCE: prefer the compact loess file ----
    if TRACE_LOESS_COMPACT.exists():
        print(f"[INFO] Using trace loess: reuse compact loess file (fast) | compact: {TRACE_LOESS_COMPACT}")
        age_tr, ghg_tr, orb_tr, ice_tr, fwf_tr, time_col, time_rule, cols_map = read_trace_loess(TRACE_LOESS_COMPACT)
    else:
        if not TRACE_RAW_FILE.exists():
            raise FileNotFoundError(f"Raw TraCE file not found: {TRACE_RAW_FILE}")
        print(f"[INFO] Compact loess not found, building from RAW with LOESS(0.2): {TRACE_RAW_FILE}")
        compact_out, time_col, time_rule, cols_map = build_compact_loess_from_raw(
            TRACE_RAW_FILE, TRACE_LOESS_COMPACT, frac=0.2
        )
        print(f"[INFO] Built compact loess: {compact_out}")
        age_tr, ghg_tr, orb_tr, ice_tr, fwf_tr, time_col, time_rule, cols_map = read_trace_loess(compact_out)

    print(f"[INFO] Time column used: {time_col}")
    print(f"[INFO] Time conversion: {time_rule}")
    print(f"[INFO] Detected loess cols: {cols_map}")

    # ---- Use only the 0–11.2 ka SN time axis as the common axis ----
    age = sn_age.copy()

    # ---- First identify the two turning-point boundaries from SN automatically ----
    a1 = window_argextreme(age, sn_z, WIN_LATE_MIN, mode="min")   # 4 ka low
    a2 = window_argextreme(age, sn_z, WIN_MID_MAX,  mode="max")   # 6-7 ka high

    if a1 is None or a2 is None or (a2 <= a1):
        raise RuntimeError(f"Unable to identify valid boundaries from the SN windows: a1={a1}, a2={a2}")

    # ---- Anchor ages (used to constrain peak-trough shape)----
    a_late_max  = window_argextreme(age, sn_z, WIN_LATE_MAX,  mode="max")
    a_late_min  = a1
    a_mid_max   = a2
    a_early_min = window_argextreme(age, sn_z, WIN_EARLY_MIN, mode="min")

    if any(v is None for v in [a_late_max, a_early_min]):
        raise RuntimeError("Unable to identify the late_max or early_min anchor. Check SN coverage.")

    # ---- Scan the parameter grid ----
    lags  = np.arange(LAG_RANGE[0], LAG_RANGE[1] + 1e-12, LAG_STEP)
    worbs = np.arange(ORB_RANGE[0], ORB_RANGE[1] + 1e-12, ORB_STEP)

    print(f"[INFO] Searching shared-lag: {len(lags)} lags ({LAG_RANGE[0]}..{LAG_RANGE[1]} step {LAG_STEP})")
    print(f"[INFO] ORB scan: {len(worbs)} values ({ORB_RANGE[0]}..{ORB_RANGE[1]} step {ORB_STEP})")
    print(f"[INFO] Stage boundaries: a1(low@3.5-4.5)={a1:.3f} ka | a2(max@6-7)={a2:.3f} ka")
    print(f"[INFO] Anchors: late_max@{a_late_max:.3f} | late_min@{a_late_min:.3f} | mid_max@{a_mid_max:.3f} | early_min@{a_early_min:.3f}")

    # ---- Estimate runtime by sampling a few optimizations ----
    pilot_pairs = []
    for il in [0, min(1, len(lags)-1)]:
        for iw in [0, min(10, len(worbs)-1)]:
            pilot_pairs.append((lags[il], worbs[iw]))
    pilot_pairs = pilot_pairs[:4]

    t0 = time.time()
    for lag, worb in pilot_pairs:
        # lagged -> interp to SN timeline
        G = interp_shift(age_tr, ghg_tr, age, lag)
        O = interp_shift(age_tr, orb_tr, age, lag)
        I = interp_shift(age_tr, ice_tr, age, lag)
        F = interp_shift(age_tr, fwf_tr, age, lag)
        Gz, Oz, Iz, Fz = zscore_nan(G), zscore_nan(O), zscore_nan(I), zscore_nan(F)
        _ = optimize_for_fixed_worb_lag(age, sn_z, Gz, Oz, Iz, Fz, worb, a1, a2,
                                        a_late_max, a_late_min, a_mid_max, a_early_min, x0=None)
    t1 = time.time()
    avg_one = (t1 - t0) / max(len(pilot_pairs), 1)
    est_total = avg_one * len(lags) * len(worbs)
    print(f"[TIME] avg per (lag,wORB) ≈ {avg_one:.3f}s | estimated total ≈ {est_total:.1f}s")

    # ---- Main search: for each wORB, find the best lag + three-stage weights ----
    scan_rows = []
    best_per_worb = []

    start = time.time()

    for worb in worbs:
        remain = 1.0 - worb
        if remain < 3*MIN_W:
            continue

        best_w = None
        x0_prev = None

        for lag in lags:
            # 1) Apply lag and interpolate onto the SN time axis
            G = interp_shift(age_tr, ghg_tr, age, lag)
            O = interp_shift(age_tr, orb_tr, age, lag)
            I = interp_shift(age_tr, ice_tr, age, lag)
            F = interp_shift(age_tr, fwf_tr, age, lag)

            # 2) Standardize (z-score)
            Gz, Oz, Iz, Fz = zscore_nan(G), zscore_nan(O), zscore_nan(I), zscore_nan(F)

            # 3) Optimize three-stage weights (ORB fixed)
            res = optimize_for_fixed_worb_lag(age, sn_z, Gz, Oz, Iz, Fz, worb, a1, a2,
                                              a_late_max, a_late_min, a_mid_max, a_early_min,
                                              x0=x0_prev)

            if res is None:
                continue

            # Warm start: use this result as the initial guess for the next lag
            x0_prev = np.array(res["w_stage"], dtype=float)

            rec = {
                "wORB": float(worb),
                "lag": float(lag),
                "loss": float(res["loss"]),
                "corr_z": float(res["corr"]),
                "w_stage": res["w_stage"],
                "mix_z": res["mix_z"]
            }

            if (best_w is None) or (rec["loss"] < best_w["loss"]):
                best_w = rec

        if best_w is not None:
            best_per_worb.append(best_w)
            scan_rows.append([best_w["wORB"], best_w["lag"], best_w["loss"], best_w["corr_z"]])

    if len(best_per_worb) == 0:
        raise RuntimeError("Scan failed: no feasible solution was found (check column detection, time range, or weight lower bounds).")

    # ---- 1% rule: within the best loss * 1.01, choose the largest wORB ----
    best_global = min(best_per_worb, key=lambda r: r["loss"])
    thr = best_global["loss"] * (1.0 + EPS_1PCT)
    cands = [r for r in best_per_worb if r["loss"] <= thr]
    chosen = max(cands, key=lambda r: r["wORB"])

    end = time.time()

    # ---- Recompute the four forcing Z series using the chosen solution (for output)----
    lag_best = chosen["lag"]
    worb_best = chosen["wORB"]
    w_stage = chosen["w_stage"]

    G = interp_shift(age_tr, ghg_tr, age, lag_best)
    O = interp_shift(age_tr, orb_tr, age, lag_best)
    I = interp_shift(age_tr, ice_tr, age, lag_best)
    F = interp_shift(age_tr, fwf_tr, age, lag_best)
    Gz, Oz, Iz, Fz = zscore_nan(G), zscore_nan(O), zscore_nan(I), zscore_nan(F)

    mix_z = chosen["mix_z"]
    mix_scaled = sn_mean + sn_std * mix_z
    residual = sn_val - mix_scaled

    # ---- Write Excel ----
    out_path = safe_out_path(OUT_XLSX)

    wb = Workbook()

    # meta sheet
    ws_meta = wb.active
    ws_meta.title = "meta"

    ws_meta.append(["SN file", str(SN_FILE)])
    ws_meta.append(["Trace loess compact", str(TRACE_LOESS_COMPACT if TRACE_LOESS_COMPACT.exists() else "(built from RAW)")])
    ws_meta.append(["Trace raw file", str(TRACE_RAW_FILE)])
    ws_meta.append(["Time column used", str(time_col)])
    ws_meta.append(["Time conversion", str(time_rule)])
    ws_meta.append(["Age range (ka BP)", f"{AGE_MIN} .. {AGE_MAX}"])
    ws_meta.append(["Stage boundaries from SN", f"a1(low@3.5-4.5)={a1:.3f}, a2(max@6-7)={a2:.3f}"])
    ws_meta.append(["Anchor windows",
                    f"late_max={WIN_LATE_MAX}, late_min={WIN_LATE_MIN}, mid_max={WIN_MID_MAX}, early_min={WIN_EARLY_MIN}"])
    ws_meta.append(["ORB scan", f"{ORB_RANGE[0]}..{ORB_RANGE[1]} step={ORB_STEP}"])
    ws_meta.append(["Lag scan", f"{LAG_RANGE[0]}..{LAG_RANGE[1]} step={LAG_STEP} (shared-lag)"])
    ws_meta.append(["Weight bounds", f"w_i >= {MIN_W} ; per stage sum = 1-wORB"])
    ws_meta.append(["Loss weights", f"ALPHA_ANCHOR={ALPHA_ANCHOR}, BETA_PEAK={BETA_PEAK}, LAMBDA_CONT={LAMBDA_CONT}, GAMMA_SMOOTH={GAMMA_SMOOTH}"])
    ws_meta.append(["ORB selection rule", f"1% rule: among loss <= best*(1+{EPS_1PCT}), choose max wORB"])
    ws_meta.append(["Chosen wORB", f"{worb_best:.3f}"])
    ws_meta.append(["Chosen lag (ka)", f"{lag_best:.3f}"])
    ws_meta.append(["Chosen loss", f"{chosen['loss']:.6f}"])
    ws_meta.append(["Chosen Corr(z)", f"{chosen['corr_z']:.4f}"])
    ws_meta.append(["SN scaling", f"MIX_scaled_to_SN = SN_mean + SN_std * MIX_z  | SN_mean={sn_mean:.6f}, SN_std={sn_std:.6f}"])
    ws_meta.append(["Residual", "Residual_SN_minus_MIX = SN - MIX_scaled_to_SN"])
    ws_meta.append(["Lag meaning", "TraCE series shifted toward older ages by +lag (response lag), then interpolated to SN timeline"])
    ws_meta.append(["Runtime (s)", f"{(end-start):.2f}"])

    # weights sheet
    ws_w = wb.create_sheet("weights")
    ws_w.append(["Stage", "w_GHG", "w_ORB", "w_ICE", "w_FWF", "sum"])
    remain = 1.0 - worb_best
    wG1,wI1,wF1, wG2,wI2,wF2, wG3,wI3,wF3 = w_stage
    ws_w.append(["Stage1 (0..a1)", f"{wG1:.3f}", f"{worb_best:.3f}", f"{wI1:.3f}", f"{wF1:.3f}", f"{(wG1+worb_best+wI1+wF1):.3f}"])
    ws_w.append(["Stage2 (a1..a2)", f"{wG2:.3f}", f"{worb_best:.3f}", f"{wI2:.3f}", f"{wF2:.3f}", f"{(wG2+worb_best+wI2+wF2):.3f}"])
    ws_w.append(["Stage3 (a2..11.2)", f"{wG3:.3f}", f"{worb_best:.3f}", f"{wI3:.3f}", f"{wF3:.3f}", f"{(wG3+worb_best+wI3+wF3):.3f}"])

    # scan sheet
    ws_scan = wb.create_sheet("ORB_scan_bestLag")
    ws_scan.append(["w_ORB", "best_lag(ka)", "best_loss", "best_corr_z"])
    for r in scan_rows:
        ws_scan.append(r)

    # Series sheet (the useful columns you wanted)
    ws_s = wb.create_sheet("series")
    df_out = pd.DataFrame({
        "Age_kaBP": age,
        "SN": sn_val,
        "SN_z": sn_z,
        "GHG_z(lagged)": Gz,
        "ORB_z(lagged)": Oz,
        "ICE_z(lagged)": Iz,
        "FWF_z(lagged)": Fz,
        "MIX4_z(piecewise)": mix_z,
        "MIX4_scaled_to_SN": mix_scaled,
        "Residual_SN_minus_MIX": residual,
    })

    ws_s.append(list(df_out.columns))
    for row in dataframe_to_rows(df_out, index=False, header=False):
        ws_s.append(row)

    wb.save(out_path)

    print(f"[DONE] Saved: {out_path}")
    print(f"[DONE] Chosen wORB={worb_best:.3f} | lag={lag_best:.3f} ka | Corr_z={chosen['corr_z']:.3f}")
    print(f"[DONE] Stage weights (GHG,ICE,FWF):")
    print(f"       Stage1: ({wG1:.3f}, {wI1:.3f}, {wF1:.3f})  | ORB={worb_best:.3f}")
    print(f"       Stage2: ({wG2:.3f}, {wI2:.3f}, {wF2:.3f})  | ORB={worb_best:.3f}")
    print(f"       Stage3: ({wG3:.3f}, {wI3:.3f}, {wF3:.3f})  | ORB={worb_best:.3f}")
    print(f"[DONE] Total time: {(end-start):.2f}s")


if __name__ == "__main__":
    main()
